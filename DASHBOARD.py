import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import openpyxl
import os
import requests
import base64
import csv
import PIL

Waitlist_latestdf = 'DATA/PROCESSED DATA/PUBLIC HOUSING/Waitlist_trend_latest.csv'
Waitlist_trend_longdf = 'DATA/PROCESSED DATA/PUBLIC HOUSING/Waitlist_trend_long.csv'
newWaitlistData = 'DATA/SOURCE DATA/Public housing/Waitlist_trend.csv'
Waitlist_breakdownsdf = 'DATA/SOURCE DATA/Public housing/Waitlist_breakdowns.csv'
ROGSSectordf = 'DATA/PROCESSED DATA/ROGS/ROGS G.csv'
ROGSHomelessnessdf = 'DATA/PROCESSED DATA/ROGS/ROGS G19.csv'
ROGSHousingdf = 'DATA/SOURCE DATA/ROGS and SHS/ROGS G18.csv'
SHSReasonsdf = 'DATA/PROCESSED DATA/SHS/Long_Form/SHS_Reasons_Long_Form.csv'
SHSClientGroupsdf = 'DATA/PROCESSED DATA/SHS/SHS_Client_Groups.csv'
PopulationStateSexAge65df = 'DATA/PROCESSED DATA/Population/Population_State_Sex_Age_to_65+.csv'
PopulationStateMonthlydf = 'DATA/PROCESSED DATA/Population/Population_State_Total_monthly.csv'
AirbnbWATotaldf = 'DATA/PROCESSED DATA/Airbnb/Airbnb_WAtotals.csv'
AirbnbGeodf = 'DATA/PROCESSED DATA/Airbnb/Airbnb_allgeo.csv'
PopulationNewFile = 'DATA/SOURCE DATA/Population/Population.csv'
updatelogfile = 'DATA/SOURCE DATA/update_log.xlsx'
SimpleWaitlistData = 'DATA/PROCESSED DATA/PUBLIC HOUSING/Waitlist_trend.csv'

waitlist_sourceURL = "https://www.parliament.wa.gov.au/Parliament/Pquest.nsf/(SearchResultsAllDesc)?SearchView&Query=(Housing%20waitlist)&Start=1&SearchOrder=4&SearchMax=1000"
waitlist_sourceText = "Parliamentary Questions"

ROGSHomelessnessSourceURL = "https://www.pc.gov.au/ongoing/report-on-government-services/2023/housing-and-homelessness/homelessness-services"
ROGSHomelessnessSourceText = "Report on Government Services 2024, Part G, Section 19 - Homelessness Services"
ROGSHousingSourceURL = "https://www.pc.gov.au/ongoing/report-on-government-services/2023/housing-and-homelessness/housing"
ROGSHousingSourceText = "Report on Government Services 2024, Part G, Section 18 - Housing"
ROGSSectorSourceURL = "https://www.pc.gov.au/research/ongoing/report-on-government-services/2022/housing-and-homelessness"
ROGSSectorSourceText = "Report on Government Services 2024, Part G - Housing and Homelessness"
SHSSourceURL = "https://www.aihw.gov.au/reports/homelessness-services/specialist-homelessness-services-monthly-data/data"
SHSSourceText = "Australian Institute of Health and Welfare - Specialist homelessness services, monthly data"

def home():
    st.set_page_config(layout="wide")
    goto = st.sidebar.selectbox('Select page', ['Waitlist', 'ROGS', 'SHS monthly data', 'Airbnb', 'Census', 'External content', 'Upload data or external content'], index=0)
    if goto == 'Waitlist':
        Waitlist_select = st.sidebar.selectbox('Select view', ['Latest data', 'Overall trend', 'Breakdowns'])
        if Waitlist_select == 'Latest data':
            waitlist_latest()
        elif Waitlist_select == 'Overall trend':
            waitlist_trendcharts()
        elif Waitlist_select == 'Breakdowns':
            waitlist_breakdowns()  
    elif goto == 'External content':
        external_resources()
    elif goto == 'ROGS':
        ROGS_select = st.sidebar.selectbox('Select ROGS page', ['Sector overview', 'Housing', 'Homelessness'])
        if ROGS_select == 'Sector overview':
            ROGS_sector()
        elif ROGS_select == 'Housing':
            ROGS_housing()
        elif ROGS_select == 'Homelessness':
            ROGS_homelessness()
    elif goto == 'SHS monthly data':
        SHS_select = st.sidebar.selectbox('Select SHS page', ['Client groups', 'Reasons for seeking assistance'])
        if SHS_select == 'Client groups':
            SHS_client_groups()
        elif SHS_select == 'Reasons for seeking assistance':
            SHS_reasons()
    elif goto == 'Airbnb':
        Airbnb_select = st.sidebar.selectbox('Select Airbnb page', ['WA total - by room type', 'Geographic filters'])
        if Airbnb_select == 'WA total - by room type':
            airbnb_wa()
        elif Airbnb_select == 'Geographic filters':
            airbnb_geo()
    elif goto == 'Upload data or external content':
        upload_data()
        show_update_log()
    elif goto == 'Census':
        census_data = st.sidebar.selectbox('Select dataset', ['Total by state', 'Geographic breakdown', 'Aboriginal and Torres Strait Islander status'])
        census(census_data)
    return

def data_updates():
    #Quick setup approach - will use API call to get latest ABS population data on or after the 21st of each quarter ending month.
    if pd.to_datetime('today').month == 3 | 6 | 9 | 12 and pd.to_datetime('today').day >= 21:
        import_population_data()
    return

def waitlist_latest():
  class WaitlistUpdate:
      def __init__(self, Date, Category, Subcategory, Metric, MetricDetail, MetricAs, MetricCalc, MetricCalcAs, Estimate, Value, FontColor):
          self.Date = Date
          self.Category = Category
          self.Subcategory = Subcategory
          self.Metric = Metric
          self.MetricDetail = MetricDetail
          self.MetricAs = MetricAs
          self.MetricCalc = MetricCalc
          self.MetricCalcAs = MetricCalcAs
          self.Estimate = Estimate
          self.Value = Value
          self.FontColor = FontColor

  waitlist_updates = []

#do not remove index from statement - necessary for tuples
  Waitlist_trend_latestdf = pd.read_csv(Waitlist_latestdf)
  for index, row in Waitlist_trend_latestdf.iterrows():
      update = WaitlistUpdate(
          Date = row['Date'],
          Category = row['Description1'],
          Subcategory = row['Description2'],
          Metric = row['Description3'],
          MetricDetail = row['Description4'],
          MetricAs = row['Description5'],
          MetricCalc = row['Description6'],
          MetricCalcAs = row['Description7'],
          Estimate = row['Estimate'],
          Value = row['Value'],
          FontColor = "red" if row['Value'] > 0 else "green"
      )
      waitlist_updates.append(update)

  TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals, ProportionPriorityApplications, ProportionPriorityIndividuals, AveragePersonsTotal, AveragePersonsPriority, AveragePersonsNonpriority = {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}

  categories = [(TotalApplications, 'Total', 'Applications'), 
                (TotalIndividuals, 'Total', 'Individuals'), 
                (PriorityApplications, 'Priority', 'Applications'), 
                (PriorityIndividuals, 'Priority', 'Individuals'),
                  (NonpriorityApplications, 'Nonpriority', 'Applications'),
                  (NonpriorityIndividuals, 'Nonpriority', 'Individuals'),
                (ProportionPriorityApplications, 'Proportion Priority', 'Applications'),
                (ProportionPriorityIndividuals, 'Proportion Priority', 'Individuals'),
                  (AveragePersonsTotal, 'Average Number Of Individuals Per Application', 'Total'),
                  (AveragePersonsPriority, 'Average Number Of Individuals Per Application', 'Priority'),
                  (AveragePersonsNonpriority, 'Average Number Of Individuals Per Application', 'Nonpriority')
                ]
  waitlist_calc_categories(waitlist_updates, categories)            
  latest_date = max(TotalApplications['Date'], TotalIndividuals['Date'], PriorityApplications['Date'], PriorityIndividuals['Date'])
  latest_date = pd.to_datetime(latest_date)
  latest_date = latest_date.strftime('%d %B %Y')
  st.markdown(f'Source: <a href="{waitlist_sourceURL}">{waitlist_sourceText} - last updated {latest_date} </a>', unsafe_allow_html=True)
  latest_table(latest_date, TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals, ProportionPriorityApplications, ProportionPriorityIndividuals, AveragePersonsTotal, AveragePersonsPriority, AveragePersonsNonpriority)
  prior_month_table(TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals)
  prior_year_table(TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals)
  return

def waitlist_calc_categories(waitlist_updates, categories):
    for category, cat1, cat2 in categories:
          category['Date'] = [x.Date for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2]
          category['Date'] = max(category['Date'])
          category['Value'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Prior month'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior month' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Prior month %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior month' and x.MetricAs == 'Percentage' and x.MetricCalc == '-']
          category['Prior month font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior month' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Prior month change second order'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior month' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
          category['Prior month change second order %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior month' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Percentage']
          category['Prior month change second order font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior month' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
          category['Prior year'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior year' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Prior year %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior year' and x.MetricAs == 'Percentage' and x.MetricCalc == '-']
          category['Prior year font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior year' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Prior year change second order'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior year' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
          category['Prior year change second order %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior year' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Percentage']
          category['Prior year change second order font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == 'prior year' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
          category['Rolling average'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == '12 month rolling average' and x.MetricDetail == '-' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Rolling average difference'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
          category['Rolling average difference %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Percentage' and x.MetricCalc == '-']
          category['Rolling average difference font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc =='-']
          category['Rolling average prior month difference'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
          category['Rolling average prior month difference %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Percentage']
          category['Rolling average prior month difference font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
          category['Rolling average prior year difference'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
          category['Rolling average prior year difference %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Percentage']
          category['Rolling average prior year difference font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Actual' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
          if cat2 == 'Individuals' and cat1 != 'Proportion Priority':
              category['per 10 000'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
              category['per 10 000 prior month'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
              category['per 10 000 prior month %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Percentage']
              category['per 10 000 prior month font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
              category['per 10 000 prior year'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
              category['per 10 000 prior year %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Percentage']
              category['per 10 000 prior year font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
              category['per 10 000 rolling average'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == '12 month rolling average' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
              category['per 10 000 rolling average difference'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
              category['per 10 000 rolling average difference %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'Percentage' and x.MetricCalc == '-']
              category['per 10 000 rolling average difference font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Difference' and x.MetricDetail == '12 month rolling average' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
              category['percentage of population'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == '-']
              category['percentage of population prior month'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
              category['percentage of population prior month %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Percentage']
              category['percentage of population prior month font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == 'change from prior month' and x.MetricCalcAs == 'Actual']
              category['percentage of population prior year'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
              category['percentage of population prior year %'] = [x.Value for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Percentage']
              category['percentage of population prior year font color'] = [x.FontColor for x in waitlist_updates if x.Category == cat1 and x.Subcategory == cat2 and x.Metric == 'Percentage of population' and x.MetricCalc == 'change from prior year' and x.MetricCalcAs == 'Actual']
    return category

def latest_table(latest_date, TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals, ProportionPriorityApplications, ProportionPriorityIndividuals, AveragePersonsTotal, AveragePersonsPriority, AveragePersonsNonpriority):
  st.markdown(f'### As at ' + latest_date)
  st.markdown(f'''
              <style>
              .custom-table {{
          width: 80%;
          border-collapse: separate;
      }}
      .custom-table th, .custom-table td {{
          font-family: Tahoma;
          text-align: center;
          border: none;
      }}
      .custom-table th {{
          background-color: transparent;
          border-bottom: none;
      }}
      .header-row {{
          font-weight: bold;
          background-color: transparent;
          border-bottom: 1px solid #d3d3d3;
      }}
      .data-row {{
          height: 1.2cm;
      }}
      .data-cell-total-count {{
          border-right: 3px dotted #d3d3d3;
          background-color: #ffff75;
          font-weight: bold;
          font-size: 18px;
      }}
              
      .data-cell-total {{
          border-right: 3px dotted #d3d3d3;
          background-color: transparent;
      }}
              
      .data-cell-nonpriority {{
          background-color: #f0f0f0;
      }}
      
      .data-cell-priority {{
          background-color: #f7e7e6;
      }}
              
      .data-cell-proportion {{
          background-color: #f7f5e6;
          font-style: italic;
      }}

      .pm-ta {{
              color: {TotalApplications["Prior month font color"][0]};
              }}
      .pm-pa {{
              color: {PriorityApplications["Prior month font color"][0]};
              }}
      .pm-npa {{
              color: {NonpriorityApplications["Prior month font color"][0]};
              }}

      .pm-ti {{
              color: {TotalIndividuals["Prior month font color"][0]};
              }}
      .pm-pi {{
              color: {PriorityIndividuals["Prior month font color"][0]};
              }}

      .pm-npi {{
              color: {NonpriorityIndividuals["Prior month font color"][0]};
              }}

      .pm-ppa {{
              color: {ProportionPriorityApplications["Prior month font color"][0]};
              }}

      .pm-ppi {{
              color: {ProportionPriorityIndividuals["Prior month font color"][0]};
              }}

      .pm-t10k {{
              color: {TotalIndividuals["per 10 000 prior month font color"][0]};
              }}

      .pm-p10k {{
              color: {PriorityIndividuals["per 10 000 prior month font color"][0]};
              }}

      .pm-np10k {{
              color: {NonpriorityIndividuals["per 10 000 prior month font color"][0]};
              }}

      .pm-tpop {{
              color: {TotalIndividuals["percentage of population prior month font color"][0]};
              }}

      .pm-ppop {{
              color: {PriorityIndividuals["percentage of population prior month font color"][0]};
              }}

      .pm-npop {{
              color: {NonpriorityIndividuals["percentage of population prior month font color"][0]};
              }}

      .pm-tavgppa {{
              color: {AveragePersonsTotal["Prior month font color"][0]};
              }}

      .pm-pavgppa {{
              color: {AveragePersonsPriority["Prior month font color"][0]};
              }}

      .pm-npavgppa {{
              color: {AveragePersonsNonpriority["Prior month font color"][0]};
              }}

      .py-ta {{
              color: {TotalApplications["Prior year font color"][0]};
              }}

      .py-pa {{
              color: {PriorityApplications["Prior year font color"][0]};
              }}

      .py-npa {{
              color: {NonpriorityApplications["Prior year font color"][0]};
              }}

      .py-ti {{
              color: {TotalIndividuals["Prior year font color"][0]};
              }}

      .py-pi {{
              color: {PriorityIndividuals["Prior year font color"][0]};
              }}

      .py-npi {{
              color: {NonpriorityIndividuals["Prior year font color"][0]};
              }}

      .py-ppa {{
          color: {ProportionPriorityApplications["Prior year font color"][0]};
              }}

      .py-ppi {{
          color: {ProportionPriorityIndividuals["Prior year font color"][0]};
                  }}

      .py-t10k {{
                  color: {TotalIndividuals["per 10 000 prior year font color"][0]};
                  }}

      .py-p10k {{
          color: {PriorityIndividuals["per 10 000 prior year font color"][0]};
          }}

      .py-np10k {{
      color: {NonpriorityIndividuals["per 10 000 prior year font color"][0]};
      }}

      .py-tpop {{
      color: {TotalIndividuals["percentage of population prior year font color"][0]};
      }}

      .py-ppop {{
      color: {PriorityIndividuals["percentage of population prior year font color"][0]};
      }}

      .py-npop {{
      color: {NonpriorityIndividuals["percentage of population prior year font color"][0]};
      }}

      .py-tavgppa {{
      color: {AveragePersonsTotal["Prior year font color"][0]};
      }}

      .py-pavgppa {{
      color: {AveragePersonsPriority["Prior year font color"][0]};
      }}

      .ra-ta {{
      color: {TotalApplications["Rolling average difference font color"][0]};
      }}

      .ra-pa {{
      color: {PriorityApplications["Rolling average difference font color"][0]};
      }}

      .ra-pi {{
      color: {PriorityIndividuals["Rolling average difference font color"][0]};
      }}

      .ra-ti {{
      color: {TotalIndividuals["Rolling average difference font color"][0]};
      }}

      .ra-pm-ta {{
      color: {TotalApplications["Rolling average prior month difference font color"][0]};
      }}

      .ra-pm-pa {{
      color: {PriorityApplications["Rolling average prior month difference font color"][0]};
      }}

      .ra-pm-pi {{
      color: {PriorityIndividuals["Rolling average prior month difference font color"][0]};
      }}

      .ra-pm-ti {{
      color: {TotalIndividuals["Rolling average prior month difference font color"][0]};
      }}

      .data-cell-total-count, .data-cell-priority, data-cell-nonpriority, .data-cell-total, .pm-ta, .pm-pa, .pm-npa, .pm-ti, .pm-pi, .pm-npi, .pm-tp, .pm-ip, .pm-tpp, .pm-pp, .pm-npp, .pm-tpop, .pm-ipop, .pm-npop, .pm-tavgppa, .pm-pavgppa, .pm-npavgppa, .py-ta, .py-pa, .py-npa, .py-ti, .py-pi, .py-npi, .py-tp, .py-ip, .py-tpp, .py-pp, .py-npp, .py-tpop, .py-ipop, .py-npop, .py-tavgppa, .py-pavgppa, .py-npavgppa, .ra-ta, .ra-pa, .ra-pi, .ra-ti, .ra-pm-ta, .ra-pm-pa, .ra-pm-pi, .ra-pm-ti {{
          height: 0.8cm;
          width: 0.8cm;
      }}

      .header-cell-total-count {{
          border-right: 3px dotted #d3d3d3;
          background-color: #ffff75;
          font-weight: bold;
          font-size: 18px;
      }}
              
      .header-cell-total {{
          border-right: 3px dotted #d3d3d3;
          background-color: transparent;
      }}
      
      .header-cell-priority {{
          background-color: #f7e7e6;
      }}
      
      .header-cell-proportion {{
          background-color: #f7f5e6;
              font-style: italic;
      }}
              
      .header-cell-nonpriority {{
          background-color: #f0f0f0;
      }}
              
      .header-cell-total, .header-cell-priority, .header-cell-nonpriority, .header-cell-proportion {{
          height: 1cm;
          width: 0.8cm;
          font-weight: bold;
          font-size: 14px;
      }}
      
      .header-applications {{
          background-color: #add8e6;
          font-weight: bold;
          font-size: 18px;
      }}

      .header-percent{{
          background-color: #b3b3f5;
          font-style: italic;
          font-weight: bold;
          font-size: 18px;
      }}

      .header-count {{
          background-color: #eeb3f5;
          font-style: italic;
          font-weight: bold;
          font-size: 18px;
      }}
              
      .header-persons-per-application {{
          background-color: #cafaf8;
          font-style: italic;
          font-size: 14px;
      }}
              
      .header-individuals {{
          background-color: #90ee90;
          font-weight: bold;
          font-size: 18px;
      }}
              
      .header-individuals-per-10k {{
          background-color: #f0e68c;
          font-style: italic;
          font-size: 14px;
      }}
      
      .header-individuals-percentage {{
          background-color: #ffd4b3;
          font-style: italic;
          font-size: 14px;
      }}
      
      .spacer-column {{
          width: 0.1cm; 
      }}
              </style>
  <table class="custom-table">
      <tr>
          <td colspan="8" class="header-applications">APPLICATIONS</td>
          <td class="spacer-column"></td>
          <td colspan="12" class="header-individuals">INDIVIDUALS</td>
      </tr>
      <tr>
          <td colspan="4" class="header-cell-total"></td>
          <td class="spacer-column"></td>
          <td colspan="3" class="header-persons-per-application">Avg. persons per application</td>
          <td class="spacer-column"></td>
          <td colspan="4" class ="header-cell-total"></td>
          <td class="spacer-column"></td>
          <td colspan="3" class="header-individuals-per-10k">Per 10,000 people</td>
          <td class="spacer-column"></td>
          <td colspan="3" class="header-individuals-percentage">As percent of population</td>
          </tr>
      <tr class="header-row">
          <td class="header-cell-total-count">TOTAL</td>
          <td class="header-cell-priority">Priority</td>
          <td class="header-cell-nonpriority">Non-priority</td>
          <td class="header-cell-proportion">% priority</td>
          <td class="spacer-column"></td>
          <td class="header-cell-total">Total</td>
          <td class="header-cell-priority">Priority</td>
          <td class="header-cell-nonpriority">Non-priority</td>
          <td class="spacer-column"></td>
          <td class="header-cell-total-count">TOTAL</td>
          <td class="header-cell-priority">Priority</td>
          <td class="header-cell-nonpriority">Non-priority</td>
          <td class="header-cell-proportion">% priority</td>
          <td class="spacer-column"></td>
          <td class="header-cell-total">Total</td>
          <td class="header-cell-priority">Priority</td>
          <td class="header-cell-nonpriority">Non-priority</td>
          <td class="spacer-column"></td>
          <td class="header-cell-total">Total</td>
          <td class="header-cell-priority">Priority</td>
          <td class="header-cell-nonpriority">Non-priority</td>
      </tr>
      <tr class="data-row">
          <td class="data-cell-total-count">{TotalApplications["Value"][0]:,.0f}</td>
          <td class="data-cell-priority">{PriorityApplications["Value"][0]:,.0f}</td>
          <td class="data-cell-nonpriority">{NonpriorityApplications["Value"][0]:,.0f}</td>
          <td class="data-cell-proportion">{ProportionPriorityApplications["Value"][0]:,.1f}%</td>
          <td class="spacer-column"></td>
          <td class="data-cell-total">{AveragePersonsTotal["Value"][0]:,.1f}</td>
          <td class="data-cell-priority">{AveragePersonsPriority["Value"][0]:,.1f}</td>
          <td class="data-cell-nonpriority">{AveragePersonsNonpriority["Value"][0]:,.1f}</td>
          <td class="spacer-column"></td>
          <td class="data-cell-total-count">{TotalIndividuals["Value"][0]:,.0f}</td>
          <td class="data-cell-priority">{PriorityIndividuals["Value"][0]:,.0f}</td>
          <td class="data-cell-nonpriority">{NonpriorityIndividuals["Value"][0]:,.0f}</td>
          <td class="data-cell-proportion">{ProportionPriorityIndividuals["Value"][0]:,.1f}%</td>
          <td class="spacer-column"></td>
          <td class="data-cell-total">{TotalIndividuals["per 10 000"][0]:,.0f}</td>
          <td class="data-cell-priority">{PriorityIndividuals["per 10 000"][0]:,.0f}</td>
          <td class="data-cell-nonpriority">{NonpriorityIndividuals["per 10 000"][0]:,.0f}</td>
          <td class="spacer-column"></td>
          <td class="data-cell-total">{TotalIndividuals["percentage of population"][0]:,.2f}%</td>
          <td class="data-cell-priority">{PriorityIndividuals["percentage of population"][0]:,.2f}%</td>
          <td class="data-cell-nonpriority">{NonpriorityIndividuals["percentage of population"][0]:,.2f}%</td>
      </tr>
  </table>
  ''', unsafe_allow_html=True)
  return

def prior_month_table(TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals):
  st.markdown('</br>', unsafe_allow_html=True)
  st.markdown(f'**Changes from prior month**')
  st.markdown(f'''
  <table class="custom-table">
          <tr class="header-row">
              <tr>
          <td colspan="7" class="header-percent">%</td>
          <td class="spacer-column"></td>
          <td colspan="7" class="header-count">NUMBER</td>
      </tr>
      <tr class="header-row">
              <tr>
              <td colspan="3", class="header-applications">APPLICATIONS</td>
              <td class="spacer-column"></td>
              <td colspan="3", class="header-individuals">INDIVIDUALS</td>
              <td class="spacer-column"></td>
  <td colspan="3", class="header-applications">APPLICATIONS</td>
              <td class="spacer-column"></td>
              <td colspan="3", class="header-individuals">INDIVIDUALS</td>
              </tr>
              <tr>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
              <td class="spacer-column"></td>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
              <td class="spacer-column"></td>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
              <td class="spacer-column"></td>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
      <tr class="data-row">
                  <td class= "pm-ta">{TotalApplications["Prior month %"][0]:,.2f}%</td>
          <td class="pm-pa">{PriorityApplications["Prior month %"][0]:,.2f}%</td>
          <td class="pm-npa">{NonpriorityApplications["Prior month %"][0]:,.2f}%</td>
          <td class="spacer-column"></td>
          <td class="pm-ti">{TotalIndividuals["Prior month %"][0]:,.2f}%</td>
          <td class="pm-pi">{PriorityIndividuals["Prior month %"][0]:,.2f}%</td>
          <td class="pm-npi">{NonpriorityIndividuals["Prior month %"][0]:,.2f}%</td>
          <td class="spacer-column"></td>
              <td class= "pm-ta">{TotalApplications["Prior month"][0]:,.0f}</td>
          <td class="pm-pa">{PriorityApplications["Prior month"][0]:,.0f}</td>
          <td class="pm-npa">{NonpriorityApplications["Prior month"][0]:,.0f}</td>
          <td class="spacer-column"></td>
          <td class="pm-ti">{TotalIndividuals["Prior month"][0]:,.0f}</td>
          <td class="pm-pi">{PriorityIndividuals["Prior month"][0]:,.0f}</td>
          <td class="pm-npi">{NonpriorityIndividuals["Prior month"][0]:,.0f}</td>
  ''', unsafe_allow_html=True)
  st.markdown('</br>', unsafe_allow_html=True)
  return

def prior_year_table(TotalApplications, TotalIndividuals, PriorityApplications, PriorityIndividuals, NonpriorityApplications, NonpriorityIndividuals):
  st.markdown(f'**Changes from prior year**')
  st.markdown(f'''
      <table class="custom-table">
          <tr class="header-row">
              <tr>
          <td colspan="7" class="header-percent">%</td>
          <td class="spacer-column"></td>
          <td colspan="7" class="header-count">NUMBER</td>
      </tr>
      <tr class="header-row">
              <tr>
              <td colspan="3", class="header-applications">APPLICATIONS</td>
              <td class="spacer-column"></td>
              <td colspan="3", class="header-individuals">INDIVIDUALS</td>
              <td class="spacer-column"></td>
  <td colspan="3", class="header-applications">APPLICATIONS</td>
              <td class="spacer-column"></td>
              <td colspan="3", class="header-individuals">INDIVIDUALS</td>
              </tr>
              <tr>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
              <td class="spacer-column"></td>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
              <td class="spacer-column"></td>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
              <td class="spacer-column"></td>
              <td class="header-cell-total-count">TOTAL</td>
              <td class="header-cell-priority">Priority</td>
              <td class="header-cell-nonpriority">Non-priority</td>
      <tr class="data-row">
                  <td class= "py-ta">{TotalApplications["Prior year %"][0]:,.2f}%</td>
          <td class="py-pa">{PriorityApplications["Prior year %"][0]:,.2f}%</td>
          <td class="py-npa">{NonpriorityApplications["Prior year %"][0]:,.2f}%</td>
          <td class="spacer-column"></td>
          <td class="py-ti">{TotalIndividuals["Prior year %"][0]:,.2f}%</td>
          <td class="py-pi">{PriorityIndividuals["Prior year %"][0]:,.2f}%</td>
          <td class="py-npi">{NonpriorityIndividuals["Prior year %"][0]:,.2f}%</td>
          <td class="spacer-column"></td>
              <td class= "py-ta">{TotalApplications["Prior year"][0]:,.0f}</td>
          <td class="py-pa">{PriorityApplications["Prior year"][0]:,.0f}</td>
          <td class="py-npa">{NonpriorityApplications["Prior year"][0]:,.0f}</td>
          <td class="spacer-column"></td>
          <td class="py-ti">{TotalIndividuals["Prior year"][0]:,.0f}</td>
          <td class="py-pi">{PriorityIndividuals["Prior year"][0]:,.0f}</td>
          <td class="py-npi">{NonpriorityIndividuals["Prior year"][0]:,.0f}</td>
  ''', unsafe_allow_html=True)
  return

def waitlist_trendcharts():
    data = pd.read_csv(Waitlist_trend_longdf)
    data['Date'] = pd.to_datetime(data['Date'])
    latest_date = data['Date'].max()
    latest_date = pd.to_datetime(latest_date, format='%Y-%m-%d').strftime('%B %Y')
    st.markdown(f'Source: <a href="{waitlist_sourceURL}">{waitlist_sourceText} - last updated {latest_date} </a>', unsafe_allow_html=True)
    class WaitlistTrend:
        def __init__(self, Date, Category, Subcategory, Metric, MetricDetail, MetricAs, MetricCalc, MetricCalcAs, Estimate, Value, FontColor):
            self.Date = Date
            self.Category = Category
            self.Subcategory = Subcategory
            self.Metric = Metric
            self.MetricDetail = MetricDetail
            self.MetricAs = MetricAs
            self.MetricCalc = MetricCalc
            self.MetricCalcAs = MetricCalcAs
            self.Estimate = Estimate
            self.Value = Value
            self.FontColor = FontColor

    waitlist_trend = [] 
    for index, row in data.iterrows():
        trend = WaitlistTrend(
            Date = row['Date'],
            Category = row['Description1'],
            Subcategory = row['Description2'],
            Metric = row['Description3'],
            MetricDetail = row['Description4'],
            MetricAs = row['Description5'],
            MetricCalc = row['Description6'],
            MetricCalcAs = row['Description7'],
            Estimate = row['Estimate'],
            Value = row['Value'],
            FontColor = "red" if row['Value'] > 0 else "green"

        )
        waitlist_trend.append(trend)   
    col1, col2, col3 = st.columns(3)

    with col1:
        select = st.selectbox('Category', ['Applications', 'Individuals'])

    with col2:
        if select == 'Applications':
            axis2 = st.selectbox('Second axis', ['Proportion of Waitlist that is priority', 'Average Number Of Individuals Per Application', 'None'])
        else:
            axis2 = st.selectbox('Second axis', ['per 10 000', 'Percentage of population', 'None'])

    with col3:
        st.markdown(f'</br>', unsafe_allow_html=True)
        Show_Rolling = col3.checkbox('Include 12 month rolling average line')
        graph_type = col3.radio('Display', ['Priority & total', 'Priority & non-priority'], horizontal=True)

    dates = [x.Date for x in waitlist_trend]
    dates = pd.DataFrame(columns=['Date'], data=dates)
    dates['Date'] = pd.to_datetime(dates['Date'])
    max_date = dates['Date'].max()
    if graph_type == 'Priority & non-priority':
        min_date = '2021-09-30'
        min_date = pd.to_datetime(min_date)
    else:
        min_date = dates['Date'].min()
    daterange = dates[(dates['Date'] >= min_date) & (dates['Date'] <= max_date)]
    daterange = daterange.sort_values(by=['Date'], ascending=True)
    daterange = daterange.drop_duplicates(subset=['Date'], keep='first')

    st.markdown("**Select date range:**")
    select_date_slider= st.select_slider('', options=daterange, value=(min_date, max_date), format_func=lambda x: x.strftime('%b %y'))
    startgraph, endgraph = list(select_date_slider)[0], list(select_date_slider)[1]
    waitlist_trend = [x for x in waitlist_trend if x.Date >= startgraph and x.Date <= endgraph]
    waitlist_totalapp = [x for x in waitlist_trend if x.Category == 'Total' and x.Subcategory == 'Applications' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_totalapp = pd.DataFrame.from_records([s.__dict__ for s in waitlist_totalapp])
    waitlist_priorityapp = [x for x in waitlist_trend if x.Category == 'Priority' and x.Subcategory == 'Applications' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_priorityapp = pd.DataFrame.from_records([s.__dict__ for s in waitlist_priorityapp])
    waitlist_nonpriorityapp = [x for x in waitlist_trend if x.Category == 'Nonpriority' and x.Subcategory == 'Applications' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_nonpriorityapp = pd.DataFrame.from_records([s.__dict__ for s in waitlist_nonpriorityapp])
    waitlist_totalind = [x for x in waitlist_trend if x.Category == 'Total' and x.Subcategory == 'Individuals' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_totalind = pd.DataFrame.from_records([s.__dict__ for s in waitlist_totalind])
    waitlist_priorityind = [x for x in waitlist_trend if x.Category == 'Priority' and x.Subcategory == 'Individuals' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_priorityind = pd.DataFrame.from_records([s.__dict__ for s in waitlist_priorityind])
    waitlist_nonpriorityind = [x for x in waitlist_trend if x.Category == 'Nonpriority' and x.Subcategory == 'Individuals' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_nonpriorityind = pd.DataFrame.from_records([s.__dict__ for s in waitlist_nonpriorityind])
    waitlist_proportionpriority = [x for x in waitlist_trend if x.Category == 'Proportion Priority' and x.Subcategory == 'Applications' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_proportionpriority = pd.DataFrame.from_records([s.__dict__ for s in waitlist_proportionpriority])
    waitlist_averageperapptot = [x for x in waitlist_trend if x.Category == 'Average Number Of Individuals Per Application' and x.Subcategory == 'Total' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_averageperapptot = pd.DataFrame.from_records([s.__dict__ for s in waitlist_averageperapptot])
    waitlist_averageperapppri = [x for x in waitlist_trend if x.Category == 'Average Number Of Individuals Per Application' and x.Subcategory == 'Priority' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_averageperapppri = pd.DataFrame.from_records([s.__dict__ for s in waitlist_averageperapppri])
    waitlist_averageperappnon = [x for x in waitlist_trend if x.Category == 'Average Number Of Individuals Per Application' and x.Subcategory == 'Nonpriority' and x.Metric == 'Number' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    waitlist_averageperappnon = pd.DataFrame.from_records([s.__dict__ for s in waitlist_averageperappnon])
    waitlist_per10000tot = [x for x in waitlist_trend if x.Category == 'Total' and x.Subcategory == 'Individuals' and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
    waitlist_per10000tot = pd.DataFrame.from_records([s.__dict__ for s in waitlist_per10000tot])
    waitlist_per10000pri = [x for x in waitlist_trend if x.Category == 'Priority' and x.Subcategory == 'Individuals' and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
    waitlist_per10000pri = pd.DataFrame.from_records([s.__dict__ for s in waitlist_per10000pri])
    waitlist_per10000non = [x for x in waitlist_trend if x.Category == 'Nonpriority' and x.Subcategory == 'Individuals' and x.Metric == 'Number' and x.MetricAs == 'per 10 000' and x.MetricCalc == '-']
    waitlist_per10000non = pd.DataFrame.from_records([s.__dict__ for s in waitlist_per10000non])
    waitlist_percentagetot = [x for x in waitlist_trend if x.Category == 'Total' and x.Subcategory == 'Individuals' and x.Metric == 'Percentage of population' and x.MetricCalc == '-']
    waitlist_percentagetot = pd.DataFrame.from_records([s.__dict__ for s in waitlist_percentagetot])
    waitlist_percentagepri = [x for x in waitlist_trend if x.Category == 'Priority' and x.Subcategory == 'Individuals' and x.Metric == 'Percentage of population' and x.MetricCalc == '-']
    waitlist_percentagepri = pd.DataFrame.from_records([s.__dict__ for s in waitlist_percentagepri])
    waitlist_percentagenon = [x for x in waitlist_trend if x.Category == 'Nonpriority' and x.Subcategory == 'Individuals' and x.Metric == 'Percentage of population' and x.MetricCalc == '-']
    waitlist_percentagenon = pd.DataFrame.from_records([s.__dict__ for s in waitlist_percentagenon])
    rolling_avgtotapp = [x for x in waitlist_trend if x.Category == 'Total' and x.Subcategory == 'Applications' and x.Metric == '12 month rolling average' and x.MetricDetail == '-'and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    rolling_avgtotapp = pd.DataFrame.from_records([s.__dict__ for s in rolling_avgtotapp])
    rolling_avgpriapp = [x for x in waitlist_trend if x.Category == 'Priority' and x.Subcategory == 'Applications' and x.Metric == '12 month rolling average' and x.MetricDetail == '-' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    rolling_avgpriapp = pd.DataFrame.from_records([s.__dict__ for s in rolling_avgpriapp])
    rolling_avgnonapp = [x for x in waitlist_trend if x.Category == 'Nonpriority' and x.Subcategory == 'Applications' and x.Metric == '12 month rolling average' and x.MetricDetail == '-'and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    rolling_avgnonapp = pd.DataFrame.from_records([s.__dict__ for s in rolling_avgnonapp])
    rolling_avgtotind = [x for x in waitlist_trend if x.Category == 'Total' and x.Subcategory == 'Individuals' and x.Metric == '12 month rolling average' and x.MetricDetail == '-'and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    rolling_avgtotind = pd.DataFrame.from_records([s.__dict__ for s in rolling_avgtotind])
    rolling_avgpriind = [x for x in waitlist_trend if x.Category == 'Priority' and x.Subcategory == 'Individuals' and x.Metric == '12 month rolling average' and x.MetricDetail == '-' and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    rolling_avgpriind = pd.DataFrame.from_records([s.__dict__ for s in rolling_avgpriind])
    rolling_avgnonind = [x for x in waitlist_trend if x.Category == 'Nonpriority' and x.Subcategory == 'Individuals' and x.Metric == '12 month rolling average' and x.MetricDetail == '-'and x.MetricAs == 'Actual' and x.MetricCalc == '-']
    rolling_avgnonind = pd.DataFrame.from_records([s.__dict__ for s in rolling_avgnonind])

    fig = go.Figure()
    if select == 'Applications':
        if graph_type == 'Priority & total':
            fig.add_trace(go.Scatter(x=waitlist_totalapp['Date'], y=waitlist_totalapp['Value'], mode='lines+markers', name='Total applications', fill='tonexty'))
            fig.add_trace(go.Scatter(x=waitlist_priorityapp['Date'], y=waitlist_priorityapp['Value'], mode='lines+markers', name='Priority applications', line=dict(color='red'), fill='tozeroy', fillcolor='palevioletred'))
            if axis2 == 'Average Number Of Individuals Per Application':
                fig.add_trace(go.Scatter(x=waitlist_averageperapptot['Date'], y=waitlist_averageperapptot['Value'], mode='lines', line=dict(color='navy', dash='dash', width=2), name='Avg persons -total', yaxis='y2'))
            if Show_Rolling:
                fig.add_trace(go.Scatter(x=rolling_avgtotapp['Date'], y=rolling_avgtotapp['Value'], mode='lines', line=dict(color='blue', width=2, dash='dot'), name='12 month rolling average - total'))
        else:
            fig.add_trace(go.Bar(x=waitlist_priorityapp['Date'], y=waitlist_priorityapp['Value'], name='Priority applications', marker_color='red'))
            fig.add_trace(go.Bar(x=waitlist_nonpriorityapp['Date'], y=waitlist_nonpriorityapp['Value'], name='Non-priority applications'))
            fig.add_trace(go.Scatter(x=waitlist_totalapp['Date'], y=waitlist_totalapp['Value'], mode='lines+markers', line=dict(color='black'), name='Total applications'))
            if Show_Rolling:
                fig.add_trace(go.Scatter(x=rolling_avgnonapp['Date'], y=rolling_avgnonapp['Value'], mode='lines', line=dict(color='blue', width=2, dash='dot'), name='12 month rolling average - total'))
            fig.update_layout(barmode='stack')
            if axis2 == 'Average Number Of Individuals Per Application':
                fig.add_trace(go.Scatter(x=waitlist_averageperapptot['Date'], y=waitlist_averageperappnon['Value'], mode='lines', line=dict(color='navy', dash='dash', width=2), name='Avg persons -total', yaxis='y2'))
        if Show_Rolling:
            fig.add_trace(go.Scatter(x=rolling_avgpriapp['Date'], y=rolling_avgpriapp['Value'], mode='lines', line=dict(color='maroon', width=2, dash='dot'), name='12 month rolling average - priority'))
        fig.update_layout(yaxis=dict(title='Applications'))
        if axis2 == 'Proportion of Waitlist that is priority':
            fig.add_trace(go.Scatter(x=waitlist_proportionpriority['Date'], y=waitlist_proportionpriority['Value'], mode='lines',  line=dict(color='maroon', dash='dash', width=2), name='Proportion priority',  yaxis='y2'))
            fig.update_layout(yaxis2=dict(overlaying='y', side='right', title='Proportion priority (%)'), showlegend=True, title_text='Waitlist applications and priority percentage')
        elif axis2 == 'Average Number Of Individuals Per Application':
            fig.add_trace(go.Scatter(x=waitlist_averageperapppri['Date'], y=waitlist_averageperapppri['Value'], mode='lines', line=dict(color='maroon', dash='dash', width=2), name='Avg persons - priority', yaxis='y2'))
            fig.update_layout(yaxis2=dict(overlaying='y', side='right', title='Average persons'), showlegend=True, title_text='Waitlist applications and average persons per application')
        else:
            fig.update_layout(showlegend=True, title_text='Waitlist applications')
    else:
        if graph_type == 'Priority & total':
            fig.add_trace(go.Scatter(x=waitlist_totalind['Date'], y=waitlist_totalind['Value'], mode='lines+markers', name='Total individuals', fill='tonexty'))
            fig.add_trace(go.Scatter(x=waitlist_priorityind['Date'], y=waitlist_priorityind['Value'], mode='lines+markers', line=dict(color='red'), name='Priority individuals', fill='tozeroy', fillcolor='palevioletred'))
            fig.update_layout(yaxis=dict(title='Individuals'))
            if Show_Rolling:
                fig.add_trace(go.Scatter(x=rolling_avgtotind['Date'], y=rolling_avgtotind['Value'], mode='lines', line=dict(color='royalblue', width=2, dash='dot'), name='12 month rolling average - total'))
            if axis2 == 'per 10 000':
                fig.add_trace(go.Scatter(x=waitlist_per10000tot['Date'], y=waitlist_per10000tot['Value'], mode='lines', line=dict(color='navy', width=2), name='per 10 000 - total', yaxis='y2'))
            elif axis2 == 'Percentage of population':
                fig.add_trace(go.Scatter(x=waitlist_percentagetot['Date'], y=waitlist_percentagetot['Value'], line=dict(color='navy', width=2), mode='lines+markers', name='% population - total', yaxis='y2'))
        else:
            fig.add_trace(go.Bar(x=waitlist_priorityind['Date'], y=waitlist_priorityind['Value'], name='Priority individuals', marker_color='red'))
            fig.add_trace(go.Bar(x=waitlist_nonpriorityind['Date'], y=waitlist_nonpriorityind['Value'], name='Non-priority individuals'))
            fig.add_trace(go.Scatter(x=waitlist_totalind['Date'], y=waitlist_totalind['Value'], mode='lines+markers', line=dict(color='black'), name='Total individuals'))
            fig.update_layout(barmode='stack')
            if Show_Rolling:
                fig.add_trace(go.Scatter(x=rolling_avgnonind['Date'], y=rolling_avgnonind['Value'], mode='lines', line=dict(color='royalblue', width=2, dash='dot'), name='12 month rolling average - total'))
            if axis2 == 'per 10 000':
                fig.add_trace(go.Scatter(x=waitlist_per10000non['Date'], y=waitlist_per10000non['Value'], mode='lines', line=dict(color='navy', dash='dash', width=2), name='per 10 000 - total', yaxis='y2'))
            elif axis2 == 'Percentage of population':
                fig.add_trace(go.Scatter(x=waitlist_percentagenon['Date'], y=waitlist_percentagenon['Value'], mode='lines', line=dict(color='navy', dash='dash', width=2), name='% population - total', yaxis='y2'))
        fig.update_layout(yaxis=dict(title='Individuals'))
        if Show_Rolling:
            fig.add_trace(go.Scatter(x=rolling_avgpriind['Date'], y=rolling_avgpriind['Value'], mode='lines', line=dict(color='maroon', width=2, dash='dot'), name='12 month rolling average - priority'))
        if axis2 == 'per 10 000':
            fig.add_trace(go.Scatter(x=waitlist_per10000pri['Date'], y=waitlist_per10000pri['Value'], mode='lines', line=dict(color='maroon', dash='dash', width=2), name='per 10 000 - priority', yaxis='y2'))
            fig.update_layout(yaxis2=dict(overlaying='y', side='right', title='per 10 000 residents'), showlegend=True, title_text='Waitlist individuals and rate per 10 000 residents')
        elif axis2 == 'Percentage of population':
            fig.add_trace(go.Scatter(x=waitlist_percentagepri['Date'], y=waitlist_percentagepri['Value'], mode='lines', line=dict(color='maroon', dash='dash', width=2), name='% population - priority', yaxis='y2'))
            fig.update_layout(yaxis2=dict(overlaying='y', side='right', title='% population'), showlegend=True, title_text='Waitlist individuals and percentage of population')
        else:
            fig.update_layout(showlegend=True, title_text='Waitlist individuals')

    fig.update_layout(
        xaxis=dict(
            tickformat="%b %y",  
            tick0=waitlist_totalapp['Date'].min(),  
            dtick="M3"
        ),
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=1.1  
        ),
    )

    col1, col2, col3 = st.columns(3)
    with col3:
        st.markdown('<table style="background-color: yellow; font-weight: bold; font-style: italic"><tr><td>Series can be toggled on/off by clicking on the legend</td></tr></table>', unsafe_allow_html=True)

    st.plotly_chart(fig, use_container_width=True)
    return

def waitlist_breakdowns():
    data = pd.read_csv(Waitlist_breakdownsdf)
    st.markdown(f'Source: <a href="{waitlist_sourceURL}">{waitlist_sourceText}</a>', unsafe_allow_html=True)

    data = data[(data['Item'] == 'Dwelling need') | (data['Item'] == 'New tenancies by region') | (data['Item'] == 'Waiting time by region') | (data['Item'] == 'Waiting time by dwelling need')]

    if 'Date' in data.columns:
        data['Date'] = pd.to_datetime(data['Date'], dayfirst=True)

    col1, col2 = st.columns(2)
    with col1:
        view = st.selectbox('Dataset', data['Item'].unique())
    with col2:
        filtered_data = data[data['Item'] == view]
        categories = filtered_data['Category'].unique()     
        categories = ['All'] + list(categories)
        category = st.selectbox('Category', categories)
    with col1:
        if category != 'All':
            filtered_data = filtered_data[filtered_data['Category'] == category]
        subcategories = filtered_data['Subcategory'].unique()
        if len(filtered_data['Subcategory'].unique()) > 1:
            subcategory = st.selectbox('Subcategory', subcategories)
        else:
            subcategory = filtered_data['Subcategory'].unique()[0]
    with col2:
        filtered_data = filtered_data[filtered_data['Subcategory'] == subcategory]
        if len(filtered_data['Region'].unique()) > 1:
            region = st.selectbox('Region', ['All'] + list(filtered_data['Region'].unique()), index=0) 
            if region != 'All':    
                filtered_data = filtered_data[filtered_data['Region'] == region]

    latest_date = filtered_data['Date'].max()
    latest_date = latest_date.strftime('%d %B %Y')
    with col2:
        st.markdown('<table style="background-color: yellow; font-weight: bold; font-style: italic"><tr><td>Series can be toggled on/off by clicking on the legend</td></tr></table>', unsafe_allow_html=True)


    if view == 'Dwelling need':
        datalabels = st.radio('Data labels on bars', ['On', 'Off'], index=1, key='datalabels', horizontal=True)
        if category == 'All':
            dwellingdata = data[data['Item'] == 'Dwelling need']
            categories = dwellingdata['Category'].unique()
            for category in categories:
                st.markdown('**{view} for {category} {subcategory} at {latest_date}**'.format(view=view, category=category, subcategory=subcategory, latest_date=latest_date), unsafe_allow_html=True)
                pie1 = filtered_data[filtered_data['Date'] == latest_date]
                piecat = pie1[pie1['Category'] == category]
                fig = px.pie(piecat, values='Value', names='Detail')
                if datalabels == 'On':
                    fig.update_traces(texttemplate='%{value:,.0f} (%{percent})', textposition='inside')
                st.plotly_chart(fig)
            for category in categories:
                st.markdown(f'**Dwelling demand by {category} over time**', unsafe_allow_html=True)
                fig2cat = filtered_data[filtered_data['Category'] == category]
                fig2 = go.Figure()
                for Detail in filtered_data['Detail'].unique():
                    fig2filtered_data = fig2cat[fig2cat['Detail'] == Detail]
                    fig2filtered_data['Date'] = fig2filtered_data['Date'].dt.strftime('%d %B %Y')
                    fig2.add_trace(go.Bar(x=fig2filtered_data['Date'], y=fig2filtered_data['Value'], name=Detail))
                if datalabels == 'On':
                    fig2.update_traces(texttemplate='%{y:.0f}', textposition='inside')
                #barmode stack
                fig2.update_layout(barmode='stack', yaxis=dict(title=f'{subcategory}'))
                st.plotly_chart(fig2, use_container_width=True)
            for category in categories:
                st.markdown(f'**Dwelling types needed by {category} - point in time comparison**', unsafe_allow_html=True)
                fig3 = go.Figure()
                cat = filtered_data[filtered_data['Category'] == category]
                dates = cat['Date'].unique()
                for date in dates:
                    fig3filtered_data = cat[cat['Date'] == date]
                    date = date.strftime('%d %B %Y')
                    fig3.add_trace(go.Bar(x=fig3filtered_data['Detail'], y=fig3filtered_data['Value'], name=date))
                if datalabels == 'On':
                    fig3.update_traces(texttemplate='%{y:.0f}', textposition='inside')
                fig3.update_layout(yaxis=dict(title=f'{subcategory}'))
                st.plotly_chart(fig3)
        else:
            st.markdown('**{view} for {category} {subcategory} at {latest_date}**'.format(view=view, category=category, subcategory=subcategory, latest_date=latest_date), unsafe_allow_html=True)
            filtered_data = filtered_data[filtered_data['Date'] == latest_date]
            fig = px.pie(filtered_data, values='Value', names='Detail')
            st.plotly_chart(fig)
            fig2 = go.Figure()
            for Detail in filtered_data['Detail'].unique():
                fig2filtered_data = filtered_data[filtered_data['Detail'] == Detail]
                fig2filtered_data['Date'] = fig2filtered_data['Date'].dt.strftime('%d %B %Y')
                fig2.add_trace(go.Bar(x=fig2filtered_data['Date'], y=fig2filtered_data['Value'], name=Detail))
            if datalabels == 'On':
                fig2.update_traces(texttemplate='%{y:.0f}', textposition='inside')
            fig2.update_layout(barmode='stack')

            st.plotly_chart(fig2, use_container_width=True)

            fig3 = go.Figure()
            dates = filtered_data['Date'].unique()
            for date in dates:
                fig3filtered_data = filtered_data[filtered_data['Date'] == date]
                date = date.strftime('%d %B %Y')
                fig3.add_trace(go.Bar(x=fig3filtered_data['Detail'], y=fig3filtered_data['Value'], name=date))
            if datalabels == 'On':
                fig3.update_traces(texttemplate='%{y:.0f}', textposition='inside')
            st.plotly_chart(fig3)
        
    elif view == 'New tenancies by region':
        datalabels = st.radio('Data labels on bars', ['On', 'Off'], index=1, key='datalabels', horizontal=True)
        dates = filtered_data['Date'].unique()
        if len(dates) < 2:
            st.markdown('Single data point only')
            date = filtered_data['Date'].unique()[0]
            clean =data[data['Item'] == view]
            clean = clean.drop(columns=['Subcategory', 'Detail', 'Item', 'Newtenanciestime', 'Date'], axis=1)
            date = date.strftime('%d %B %Y')
            clean['Category'] = clean['Category'].str.contains('Priority')
            clean['Category'] = clean['Category'].replace(True, 'Priority')
            clean['Category'] = clean['Category'].replace(False, 'Total')
            clean = clean.pivot_table(index='Region', columns='Category', values='Value', aggfunc='sum')
            clean.loc['WA total'] = clean.sum()    
            clean['Priority %'] = clean['Priority'] / clean['Total'] * 100
            clean['Priority %'] = clean['Priority %'].round(1)
            region_need = Waitlist_breakdownsdf[Waitlist_breakdownsdf['Item'] == 'Region need']
            region_dates = region_need['Date'].unique()
            latest_date = region_dates.max()
            region_need = region_need[region_need['Date'] == latest_date]
            region_need['Category'] = region_need['Category'].str.contains('Priority')
            region_need['Category'] = region_need['Category'].replace(True, 'Priority')
            region_need['Category'] = region_need['Category'].replace(False, 'Total')
            region_need = region_need[region_need['Subcategory'] == 'Applications']
            region_need = region_need.drop(columns=['Subcategory', 'Detail', 'Item', 'Newtenanciestime', 'Date'], axis=1)
            #pivot table
            region_need = region_need.pivot_table(index='Region', columns='Category', values='Value', aggfunc='sum')
            #create WA total row
            region_need.loc['WA total'] = region_need.sum()
            region_need['Priority %'] = region_need['Priority'] / region_need['Total'] * 100
            #proportion priority to .1f
            region_need['Priority %'] = region_need['Priority %'].round(1)
            clean = pd.merge(clean, region_need, on='Region', suffixes=('', ' waitlist'))
            clean[f'% housed - Priority'] = clean['Priority'] / clean['Priority waitlist'] * 100
            clean[f'% housed - Priority'] = clean[f'% housed - Priority'].round(1)
            clean[f'% housed - Total'] = clean['Total'] / clean['Total waitlist'] * 100
            clean[f'% housed - Total'] = clean[f'% housed - Total'].round(1)
            regionfigdata = clean.reset_index()
            regionfig = go.Figure()
            regionfig.add_trace(go.Bar(x=regionfigdata['Region'], y=regionfigdata[f'% housed - Priority'], name=f'% housed - Priority'))
            regionfig.add_trace(go.Bar(x=regionfigdata['Region'], y=regionfigdata[f'% housed - Total'], name=f'% housed - Total'))
            regionfig.update_layout(barmode='group', yaxis=dict(title='%'), title_text = f'Percentage of waitlist at {latest_date} housed in 12months to to {date} - group by region')
            if datalabels == 'On':
                regionfig.update_traces(texttemplate='%{y:.1f}', textposition='inside')
            st.plotly_chart(regionfig)
            regionlist = list(regionfigdata['Region'].unique())
            housed = regionfigdata[['Region', '% housed - Priority', '% housed - Total']]
            housed = housed.T
            housed = housed.reset_index()
            housed.columns = housed.iloc[0]
            housed = housed.drop(0)
            housed = housed.rename(columns={'Region': 'Category'})
            regionfig2 = go.Figure()
            for region in regionlist:
                regionfig2.add_trace(go.Bar(x=housed['Category'], y=housed[region], name=region))
            regionfig2.update_layout(barmode='group', yaxis=dict(title='%'), title_text = f'Percentage of waitlist at {latest_date} housed in 12 months to {date} - group by applicant type')
            if datalabels == 'On':
                regionfig2.update_traces(texttemplate='%{y:.1f}', textposition='inside')
            st.plotly_chart(regionfig2)
            st.write(housed)
            st.write(clean)
        else:
            for region in filtered_data['Region'].unique():
                st.markdown(f'**New tenancies in {region}**', unsafe_allow_html=True)
                regionchart = go.Figure()
                region_filtered_data = filtered_data[filtered_data['Region'] == region]
                region_filtered_data['Date'] = region_filtered_data['Date'].dt.strftime('%d %B %Y')
                regionchart.add_trace(go.Bar(x=region_filtered_data['Date'], y=region_filtered_data['Value'], name=region))
                regionchart.update_layout(yaxis=dict(title='New tenancies'))
                st.plotly_chart(regionchart, use_container_width=True)

    elif view == 'Waiting time by dwelling need':
        #if Category contains Priority, change to Priority, else Total
        filtered_data['Category'] = filtered_data['Category'].str.contains('Priority')
        filtered_data['Category'] = filtered_data['Category'].replace(True, 'Priority Waitlist')
        filtered_data['Category'] = filtered_data['Category'].replace(False, 'Total Waitlist')
        #if len Subcategory >1 selectbox
        if len(filtered_data['Subcategory'].unique()) > 1:
            #selectbox Subcategory
            subcategory = st.selectbox('Metric', filtered_data['Subcategory'].unique())
            #filter data to only include selected Subcategory
            filtered_data = filtered_data[filtered_data['Subcategory'] == subcategory]
            #drop Subcategory
            filtered_data = filtered_data.drop(columns=['Subcategory'], axis=1)
        else:
            #drop Subcategory
            filtered_data = filtered_data.drop(columns=['Subcategory'], axis=1)
        #drop Item, Newtenanciestime, Region
        filtered_data = filtered_data.drop(columns=['Item', 'Newtenanciestime', 'Region'], axis=1)
        #pivot table
        #round value to .0f
        filtered_data['Value'] = filtered_data['Value'].round(0)
        #date as string
        filtered_data['Date'] = filtered_data['Date'].dt.strftime('%d %B %Y')
        dwellingwait = go.Figure()
        if len(filtered_data['Date'].unique()) ==1:
            date = filtered_data['Date'].unique()[0]
            for category in filtered_data['Category'].unique():
                categorydata = filtered_data[filtered_data['Category'] == category]
                dwellingwait.add_trace(go.Bar(x=categorydata['Detail'], y=categorydata['Value'], name=category))
            dwellingwait.update_layout(barmode='group', yaxis=dict(title='Waiting time (weeks)'), title_text = f'Waiting time by dwelling need - {subcategory} - {date}')
            st.plotly_chart(dwellingwait, use_container_width=True)
        else:
            if category != 'All':
                for date in filtered_data['Date'].unique():
                    datefiltered_data = filtered_data[filtered_data['Date'] == date]
                    dwellingwait.add_trace(go.Bar(x=datefiltered_data['Detail'], y=datefiltered_data['Value'], name=date))
                dwellingwait.update_layout(barmode='group', yaxis=dict(title='Waiting time (weeks)'), title_text = f'Waiting time by dwelling need - {subcategory} - {category}', showlegend=True)
            else:
                for cat in filtered_data['Category'].unique():
                    catwaitdwellfig = go.Figure()
                    catfiltered_data = filtered_data[filtered_data['Category'] == cat]
                    for date in catfiltered_data['Date'].unique():
                        datefiltered_data = catfiltered_data[catfiltered_data['Date'] == date]
                        catwaitdwellfig.add_trace(go.Bar(x=datefiltered_data['Detail'], y=datefiltered_data['Value'], name=date))
                    catwaitdwellfig.update_layout(barmode='group', yaxis=dict(title='Waiting time (weeks)'), title_text = f'Waiting time by dwelling need - {subcategory} - {cat}', showlegend=True)
                    st.plotly_chart(catwaitdwellfig, use_container_width=True)
        
        
        whenjoinorhouse = filtered_data.copy()
        #add column Forecast house date = today + weeks(Value)
        whenjoinorhouse['Forecast house date'] = pd.to_datetime(whenjoinorhouse['Date']) + pd.to_timedelta(whenjoinorhouse['Value'], unit='w')
        #add column Backcast apply date = today - weeks(Value)
        whenjoinorhouse['Backcast apply date'] = pd.to_datetime(whenjoinorhouse['Date']) - pd.to_timedelta(whenjoinorhouse['Value'], unit='w')
        #set columns to string 
        whenjoinorhouse['Forecast house date'] = whenjoinorhouse['Forecast house date'].dt.strftime('%d %B %Y')
        whenjoinorhouse['Backcast apply date'] = whenjoinorhouse['Backcast apply date'].dt.strftime('%d %B %Y')
        st.write(whenjoinorhouse)

    elif view == 'Waiting time by region':
        #repeat similar to above
        #if Category contains Priority, change to Priority, else Total
        filtered_data['Category'] = filtered_data['Category'].str.contains('Priority')
        filtered_data['Category'] = filtered_data['Category'].replace(True, 'Priority Waitlist')
        filtered_data['Category'] = filtered_data['Category'].replace(False, 'Total Waitlist')
        #if len Subcategory >1 selectbox
        if len(filtered_data['Subcategory'].unique()) > 1:
            #selectbox Subcategory
            subcategory = st.selectbox('Metric', filtered_data['Subcategory'].unique())
            #filter data to only include selected Subcategory
            filtered_data = filtered_data[filtered_data['Subcategory'] == subcategory]
            #drop Subcategory
            filtered_data = filtered_data.drop(columns=['Subcategory'], axis=1)
        else:
            #drop Subcategory
            filtered_data = filtered_data.drop(columns=['Subcategory'], axis=1)

        #drop Item, Newtenanciestime, Detail
        filtered_data = filtered_data.drop(columns=['Item', 'Newtenanciestime', 'Detail'], axis=1)
        #pivot table
        #round value to .0f
        filtered_data['Value'] = filtered_data['Value'].round(0)
        #forecast house date = today + weeks(Value)
        filtered_data['Forecast house date'] = pd.to_datetime(filtered_data['Date']) + pd.to_timedelta(filtered_data['Value'], unit='w')
        #backcast apply date = today - weeks(Value)
        filtered_data['Backcast apply date'] = pd.to_datetime(filtered_data['Date']) - pd.to_timedelta(filtered_data['Value'], unit='w')
        #set columns to string
        filtered_data['Date'] = filtered_data['Date'].dt.strftime('%d %B %Y')
        filtered_data['Forecast house date'] = filtered_data['Forecast house date'].dt.strftime('%d %B %Y')
        filtered_data['Backcast apply date'] = filtered_data['Backcast apply date'].dt.strftime('%d %B %Y')
        st.write(filtered_data)
        return
    
def show_update_log():
    update_log = pd.read_excel('DATA/SOURCE DATA/update_log.xlsx')
    st.write('Update Log')
    st.table(update_log)
    return

def SHS_reasons():
    df = pd.read_csv(SHSReasonsdf)
    df['MEASURE'] = df['MEASURE'].fillna('Persons')
    df = df.rename(columns={'REASON FOR SEEKING ASSISTANCE': 'REASON'})
    df_latest_date = df[df['DATE'] == df['DATE'].max()]
    latest_date = df_latest_date['DATE'].max()
    df_latest_date = df_latest_date[df_latest_date['MEASURE'] == 'Persons']
    df_latest_total = df_latest_date[df_latest_date['REASON'] == 'Total clients']
    df_latest_total = df_latest_total.drop(columns=['REASON', 'MONTH', 'GROUP', 'MEASURE', 'DATE'])
    df_latest_reasons = df_latest_date[df_latest_date['REASON'] != 'Total clients']
    df_latest_reasons = df_latest_reasons.drop(columns=['MONTH', 'GROUP', 'MEASURE', 'DATE'])
    df_latest_reasons = df_latest_reasons.merge(df_latest_total, on=['STATE'])
    df_latest_reasons = df_latest_reasons.rename(columns={'VALUE_x': 'VALUE', 'VALUE_y': 'Total clients'})
    df_latest_reasons['proportion'] = (df_latest_reasons['VALUE'] / df_latest_reasons['Total clients'])*100
    df_latest_reasons = df_latest_reasons.drop(columns=['Total clients'])
    nat_reasons = df_latest_reasons[df_latest_reasons['STATE'] == 'National']
    nat_reasons = nat_reasons.groupby('REASON').sum().reset_index().sort_values(by='proportion', ascending=False)
    top_reasons = nat_reasons['REASON'].head(3).tolist()
    wa_reasons = df_latest_reasons[df_latest_reasons['STATE'] == 'WA']
    wa_reasons = wa_reasons.groupby('REASON').sum().reset_index().sort_values(by='VALUE', ascending=False)
    top_reasons_wa = wa_reasons['REASON'].head(3).tolist()
    top_reasons = top_reasons + top_reasons_wa
    top_reasons = list(dict.fromkeys(top_reasons))
    df_latest_reasons = df_latest_reasons[df_latest_reasons['REASON'].isin(top_reasons)]
    df_latest_reasons_prop = df_latest_reasons
    df_latest_reasons_prop = df_latest_reasons_prop.drop(columns=['VALUE'])
    df_latest_reasons_count = df_latest_reasons
    df_latest_reasons_count = df_latest_reasons_count.drop(columns=['proportion'])
    df_top_proportion = df_latest_reasons_prop.pivot_table(index=['STATE'], columns='REASON', values='proportion').reset_index()
    latest_date = pd.to_datetime(latest_date, format='%Y-%m-%d').strftime('%B %Y')
    st.markdown(f'Source: <a href="{SHSSourceURL}">{SHSSourceText} - last updated {latest_date} </a>', unsafe_allow_html=True)
    states = st.multiselect('Show', ['National', 'WA', 'NSW', 'Vic', 'Qld', 'SA', 'Tas', 'NT', 'ACT'], default=['National', 'WA', 'NSW', 'Vic', 'Qld', 'SA', 'Tas', 'NT', 'ACT'])
    fig = go.Figure()
    df_top_proportion = df_top_proportion[df_top_proportion['STATE'].isin(states)]
    for reason in top_reasons:
        fig.add_trace(go.Bar(x=df_top_proportion['STATE'], y=df_top_proportion[reason], name=reason))
    fig.update_layout(barmode='group', xaxis={'categoryorder':'array', 'categoryarray': states})
    fig.update_layout(title={'text': 'Proportion of clients reporting a top reason for seeking assistance', 'x': 0.5, 'xanchor': 'center'})
    fig.update_layout(legend={'title': 'Reason for Seeking Assistance'})
    fig.update_layout(yaxis={'title': '% of clients'})
    st.plotly_chart(fig)
    return

def SHS_client_groups():
    df = pd.read_csv(SHSClientGroupsdf)
    df.columns = df.columns.str.upper()
    if 'MONTH' in df.columns:
        df['DATE'] = '20' + df['MONTH'].str[3:5] + '-' + df['MONTH'].str[0:3] + '-01'
        df['DATE'] = pd.to_datetime(df['DATE'], format='%Y-%b-%d')
        df['DATE'] = df['DATE'] + pd.offsets.MonthEnd(0)
    population = pd.read_csv(PopulationStateMonthlydf)
    #columns to upper
    population.columns = population.columns.str.upper()
    population['DATE'] = pd.to_datetime(population['DATE'], format='%d/%m/%Y', dayfirst=True)
    population = population.sort_values(by='DATE', ascending=True)
    regions = df.columns[3:12]
    df['DATE'] = pd.to_datetime(df['DATE'], format='%Y-%m-%d', errors='coerce')
    df = df.sort_values(by='DATE', ascending=True)
    latest_date = df['DATE'].max()
    latest_date = pd.to_datetime(latest_date, format='%Y-%m-%d').strftime('%B %Y')
    df_tot = df[df['SEX'] == 'Total']
    df_tot = pd.merge(df_tot, population, on='DATE', how='left')
    df_tot = df_tot.fillna(method='ffill')
    st.markdown(f'Source: <a href="{SHSSourceURL}">{SHSSourceText} - last updated {latest_date} </a>', unsafe_allow_html=True)
    per10k ={}
    for region in regions:
        region_per_10k = f'{region}_PER_10k'
        per10k[region] = region_per_10k
    propnat = {}
    for region in regions:
        region_prop_nat = f'{region}_PROPORTION_OF_NATIONAL'
        propnat[region] = region_prop_nat
    groups = df['CLIENT GROUP'].unique()
    groups = groups.tolist()
    col1, col2, col3 = st.columns(3)
    with col1:
        view = st.radio('Select view', ['Number of clients', 'Number of clients per 10,000 people'], index=0)
    if view == 'Number of clients per 10,000 people':
        groups.remove('Number of nights in short-term/emergency accommodation')
    with col2:
        region = st.selectbox('Select region', regions, index=3)
    group = st.selectbox('Select client group', groups, index=7)
    df = df[df['CLIENT GROUP'] == group]
    df_tot = df_tot[df_tot['CLIENT GROUP'] == group]
    fig = go.Figure()
    if view == 'Number of clients':
        with col3:
            sex = st.radio('Sex breakdown', ['On', 'Off'])
    else:
        sex  = 'Off'
    df_fem = df[df['SEX'] == 'Female']
    df_mal = df[df['SEX'] == 'Male']
    if view == 'Number of clients':
        xvalfem = df_fem['DATE']
        xvalmal = df_mal['DATE']
        xvaltot = df_tot['DATE']
        yvalfem = df_fem[region]
        yvalmal = df_mal[region]
        yvaltot = df_tot[region]
        if group != 'Number of nights in short-term/emergency accommodation':
            ytitle = 'Number of clients'
        else:
            ytitle = 'Number of nights'
    elif view == 'Number of clients per 10,000 people':
        ytitle = 'Number of clients per 10,000 people'
        xvaltot = df_tot['DATE']
        region_pop = f'{region}_POPULATION'
        yvaltot = df_tot[region]/df_tot[region_pop]*10000
    if sex == 'On':
        fig.add_trace(go.Bar(x=xvalfem, y=yvalfem, name='Female'))
        fig.add_trace(go.Bar(x=xvalmal, y=yvalmal, name = 'Male'))
    else:
        fig.add_trace(go.Bar(x=xvaltot, y=yvaltot))
    fig.update_layout(barmode='stack', title=f'WA - {group}', yaxis_title=ytitle)
    st.plotly_chart(fig)
    return    

def ROGS_sector():
    df = pd.read_csv(ROGSSectordf, encoding='latin-1')
    #sort year ascending
    df = df.sort_values(by='Year', ascending=True)
    df['Year'] = df['Year'].astype(str)

    st.markdown(f'Source: <a href="{ROGSSectorSourceURL}">{ROGSSectorSourceText}</a>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    Population = pd.read_csv(PopulationStateSexAge65df)
    #Population filter for All ages, Total, mm=06
    Population['Date'] = pd.to_datetime(Population['Date'], format='%d/%m/%y', dayfirst=True, errors='coerce')

    df = df.rename(columns={'Aust': 'National'})
    regions = ['National', 'WA', 'Vic', 'Qld', 'SA', 'NSW', 'Tas', 'NT', 'ACT']
    #df long = melt on regions, value_name='Value', var_name='Region'
    cols = df.columns.tolist()
    #remove regions from cols
    for region in regions:
        cols.remove(region)
    dflong = pd.melt(df, id_vars=cols, value_vars=regions, var_name='Region', value_name='Value')

    #filter out measure = Households residing in community housing
    df = df[df['Measure'] != 'Households residing in community housing']

    with col1:
        select_measure_sector = st.selectbox('Select measure', df['Measure'].unique())                
        datalabels = st.radio('Data labels', ['On', 'Off'], index=0, horizontal=True)
    with col3:
        st.markdown('<table style="background-color: yellow; font-weight: bold; font-style: italic"><tr><td>Series can be toggled on/off by clicking on the legend</td></tr></table>', unsafe_allow_html=True)

    df = df[df['Measure'] == select_measure_sector]
    df['Year'] = df['Year'].astype(str)

    if select_measure_sector == "Recurrent expenditure":
        with col2:
            regions_sector = st.multiselect('Select regions', regions, default=regions)
        ytitle = df['Unit'].unique()[0] + ' (' + df['Year_Dollars'].unique()[0] + ')'
        dfRE = df[df['Description3'] == 'Total']
        CRA = dfRE[dfRE['Description2'] == 'Commonwealth Rent Assistance (CRA)']
        NHHA = dfRE[dfRE['Description2'] == 'Total NHHA related expenditure']

        #category bar chart, x=year, y=df[region] for region in regions, color=Description1, group
        fig = go.Figure()
        for region in regions_sector:
            fig.add_trace(go.Bar(x=NHHA['Year'], y=NHHA[region], name=region))
        fig.update_layout(barmode='group', title='NHHA funding', yaxis_title=ytitle)
        #legend below chart
        fig.update_layout(legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="right", x=1))
        if datalabels == 'On':
            fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
        st.plotly_chart(fig)

        fig2 = go.Figure()
        for region in regions_sector:
            fig2.add_trace(go.Bar(x=CRA['Year'], y=CRA[region], name=region))
        fig2.update_layout(barmode='group', title='CRA funding', yaxis_title=ytitle)
        #legend below chart
        fig2.update_layout(legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="right", x=1))
        if datalabels == 'On':
            fig2.update_traces(texttemplate='%{y:.2s}', textposition='inside')
        st.plotly_chart(fig2)

    if select_measure_sector == "Low income rental households":
        with col2:
            regions_sector = st.multiselect('Select regions', regions, default=regions)
        dfLIH = df[df['Uncertainty'].isna()]
        select_year_sector = st.selectbox('Select year', dfLIH['Year'].unique())
        dfLIH = dfLIH[dfLIH['Year'] == select_year_sector]
        ytitle1 = "Proportion"
        ytitle2 = "Number"
        dfLIH = dfLIH[dfLIH['Description3'] == 'Paying more than 30% of income on housing costs']
        dfProp = dfLIH[dfLIH['Description4'] == 'Proportion']
        dfNum = dfLIH[dfLIH['Description4'] == 'Number']
        fig = go.Figure()
        fig2 = go.Figure()
        for region in regions_sector:
            fig.add_trace(go.Bar(x=dfProp['Description2'], y=dfProp[region], name=region))
            fig2.add_trace(go.Bar(x=dfNum['Description2'], y=dfNum[region], name=region))
        fig.update_layout(barmode='group', title='Proportion of low income rental households paying more than 30% of income on housing costs', xaxis_title="Remoteness", yaxis_title=ytitle1)
        fig2.update_layout(barmode='group', title='Number of low income rental households paying more than 30% of income on housing costs', xaxis_title="Remoteness", yaxis_title=ytitle2)
        if datalabels == 'On':
            fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            fig2.update_traces(texttemplate='%{y:.2s}', textposition='inside')
        st.plotly_chart(fig)
        st.plotly_chart(fig2)

    if select_measure_sector == "Housing affordability":
        ytitle = df['Description2'].unique()[0]
        charttitle = df['Description1'].unique()[0]
        compare_sector = st.radio('Compare', ['States', 'Years', 'States & years'], horizontal=True)
        if compare_sector == 'States':
            regions_sector = st.multiselect('Select regions', regions, default=regions)
            select_year_sector= st.selectbox('Select year', df['Year'].unique())
            dfHA = df[df['Year'] == select_year_sector]
            fig = go.Figure()
            for region in regions_sector:
                fig.add_trace(go.Bar(x=dfHA['Year'], y=dfHA[region], name=region))
            fig.update_layout(barmode='group', title=charttitle, xaxis_title="Year", yaxis_title=ytitle)
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)
        if compare_sector == 'Years':
            regions_sector = st.selectbox('Select region', regions)
            years_sector = st.multiselect('Select years', df['Year'].unique(), default=df['Year'].unique())
            fig = go.Figure()
            for year in years_sector:
                dfHA = df[df['Year'] == year]
                fig.add_trace(go.Bar(x=dfHA['Year'], y=dfHA[regions], name=year))
            fig.update_layout(barmode='group', title=charttitle, xaxis_title="Year", yaxis_title=ytitle)
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)
        if compare_sector == 'States & years':
            regions_sector = st.multiselect('Select regions', regions, default=regions)
            years_sector = st.multiselect('Select years', df['Year'].unique(), default=df['Year'].unique())
            dfHA = df[df['Year'].isin(years_sector)]
            dflong = pd.melt(dfHA, id_vars=cols, value_vars=regions, var_name='Region', value_name='Value')
            dflong = dflong[dflong['Region'].isin(regions)]
            #sort dflong by Year ascending
            dflong = dflong.sort_values(by=['Year'], ascending=True)

            fig = go.Figure()
            fig.add_trace(go.Bar(x=[dflong['Region'],dflong['Year']], y=dflong['Value']))
            #add figure inside bar
            fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            fig.update_layout(barmode='group', title=charttitle, yaxis_title=ytitle)
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)
    if select_measure_sector == "Housing composition by tenure type":
        df['Description4'] = df['Description4'].fillna(df['Description3'])
        compare_sector = st.radio('Compare', ['States', 'Years', 'States & years'], horizontal=True)
        df = df[df['Description4'] != 'Total renters']
        df = df[df['Uncertainty'].isna()]
        #sort year ascending
        df = df.sort_values(by=['Year'], ascending=True)
        if compare_sector == 'States':
            regions_sector = st.multiselect('Select regions', regions, default=regions)
            select_year_sector = st.selectbox('Select year', df['Year'].unique())
            df = df[df['Year'] == select_year_sector]   
            fig = go.Figure()
            for region in regions_sector:
                fig.add_trace(go.Bar(x=df['Description4'], y=df[region], name=region))
            fig.update_layout(barmode='group', title='Proportion of renters by tenure type', xaxis_title="Tenure type", yaxis_title="Proportion")
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)
        if compare_sector == 'Years':
            regions_sector = st.selectbox('Select region', regions)
            #YEAR TO STRING
            df['Year'] = df['Year'].astype(str)
            years_sector = st.multiselect('Select years', df['Year'].unique(), default=df['Year'].unique())
            fig = go.Figure()
            for year in years_sector:
                df = df[df['Year'] == year]
                fig.add_trace(go.Bar(x=df['Description4'], y=df[regions_sector], name=year))
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)
        if compare_sector == 'States & years':
            regions_sector = st.multiselect('Select regions', regions, default=regions)
            years_sector = st.multiselect('Select years', df['Year'].unique(), default=df['Year'].unique())
            tenure = st.selectbox('Select tenure type', df['Description4'].unique())
            df = df[df['Description4'] == tenure]
            df = df[df['Year'].isin(years_sector)]
            dflong = pd.melt(df, id_vars=cols, value_vars=regions, var_name='Region', value_name='Value')
            #for year in years, filter dflong for year, plotly express bar, x=Region, y=Value, color=Region, facet_col=Year
            dflong['Year'] = dflong['Year'].astype(str)
            #dflong Region in regions
            dflong = dflong[dflong['Region'].isin(regions_sector)]
            fig = go.Figure()
            dflong = dflong[dflong['Description4'] == tenure]
            fig.add_trace(go.Bar(x=[dflong['Region'], dflong['Year']], y=dflong['Value']))
            if tenure == 'Home owners without a mortgage':
                    fig.update_traces(marker_color='green')
            if tenure == 'Home owners with a mortgage':
                    fig.update_traces(marker_color='blue')
            if tenure == 'Private rental':
                    fig.update_traces(marker_color='red')
            if tenure == 'Public housing':
                    fig.update_traces(marker_color='orange')

            fig.update_layout(barmode='stack', title='Proportion of renters by tenure type', xaxis_title="Tenure type", yaxis_title="Proportion")
            fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            #legend title Tenure type
            fig.update_layout(legend_title_text='Tenure type')
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)

    if select_measure_sector == 'Income units receiving CRA':
        with col2:
            select_view = st.selectbox('Select view', ['Demographics', 'Housing affordability'])
        if select_view == 'Demographics':
            #in Description2, replace "Income unit" with "family type"
            df['Description2'] = df['Description2'].str.replace('Income unit', 'Family type')
            #replace All with Support payment type
            df['Description2'] = df['Description2'].str.replace('All', 'Support payment type')
            df['Description2'] = df['Description2'].str.replace('Aged under 25 years', 'Aged under 25 / aged over 75')
            df['Description2'] = df['Description2'].str.replace('Aged 75 years or over', 'Aged under 25 / aged over 75')
            #remove from Description2: Major cities Inner regional Outer regional Remote Very remote, Disability Support Pension recipient, Non-Indigenous, Dependent children in income units
            df= df[df['Description2'] != 'Major cities']
            df= df[df['Description2'] != 'Inner regional']
            df= df[df['Description2'] != 'Outer regional']
            df= df[df['Description2'] != 'Remote']
            df= df[df['Description2'] != 'Very remote']
            df= df[df['Description2'] != 'Disability Support Pension recipient']
            df= df[df['Description2'] != 'Non-Indigenous']
            df= df[df['Description2'] != 'Dependent children in income units']
            
            col1, col2 = st.columns(2)
            df = df[df['Description1'] == 'Income units receiving CRA']
            with col1:
                select_sector = st.selectbox('Detail', df['Description2'].unique())
                df = df[df['Description2'] == select_sector]
                if select_sector == 'Paying enough rent to be eligible for maximum assistance':
                    df = df[df['Description2'] == select_sector]
                    #allow select region, select year
                    with col1:
                        select_years_sector = st.multiselect('Select year', df['Year'].unique(), default=df['Year'].unique())
                    df = df[df['Year'].isin(select_years_sector)]
                    fig = go.Figure()
                    #for region in regions, add trace to fig
                    for region in regions:
                        fig.add_trace(go.Bar(x=df['Year'], y=df[region], name=region))
                    fig.update_layout(barmode='group', title='Paying enough rent to be eligible for maximum assistance', xaxis_title="Year", yaxis_title="%")
                    #add data labels inside bars
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig)
                if select_sector == 'Geographic location':
                    #select year
                    with col1:
                        select_years_geo = st.multiselect('Select year', df['Year'].unique(), default=df['Year'].unique())
                    with col2:
                        select_unit = st.selectbox('Select unit', df['Unit'].unique())
                    df = df[df['Year'].isin(select_years_geo)]
                    df = df[df['Unit'] == select_unit]
                    fig = go.Figure()
                    for region in regions:
                        fig.add_trace(go.Bar(x=[df['Description3'],df['Year']], y=df[region], name=region))
                    #add figure inside bar
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    fig.update_layout(barmode='group', title='By geographic location', xaxis_title="Year", yaxis_title=select_unit)
                    st.plotly_chart(fig)
                if select_sector == 'Total':
                    df = df[df['Description2'] == select_sector]
                    fig = go.Figure()
                    for region in regions:
                        fig.add_trace(go.Bar(x=df['Year'], y=df[region], name=region))
                    fig.update_layout(barmode='group', title='Total receiving CRA', xaxis_title="Year", yaxis_title=df['Unit'].unique()[0])
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig)
                if select_sector == "Aboriginal and Torres Strait Islander":
                    df = df[df['Description2'] == select_sector]
                    df1 = df[df['Table_Number'] == 'GA.8']
                    df2 = df[df['Table_Number'] == 'GA.9']
                    a8list = df1['Description3'].unique().tolist()
                    #remove total
                    a8list.remove('Total')
                    fig = go.Figure()
                    df1pc = df1[df1['Unit'] == '%']
                    df1no = df1[df1['Unit'] == 'no.']
                    for region in regions:
                        fig.add_trace(go.Bar(x=df1pc['Description3'], y=df1pc[region], name=region))
                    fig.update_layout(barmode='stack', title='Aboriginal and Torres Strait Islander recipient family types - %', xaxis_title="Family type", yaxis_title='%')
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig)
                    fig2 = go.Figure()
                    for region in regions:
                        fig2.add_trace(go.Bar(x=df1no['Description3'], y=df1no[region], name=region))
                    fig2.update_layout(barmode='stack', title='Aboriginal and Torres Strait Islander recipient family types - no.', xaxis_title="Family type", yaxis_title='no.')
                    if datalabels == 'On':
                        fig2.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig2)
                    fig3 = go.Figure()
                    for region in regions:
                        fig3.add_trace(go.Bar(x=df2['Description3'], y=df2[region], name=region))
                    fig3.update_layout(barmode='stack', title='Aboriginal and Torres Strait Islander recipient payment types - %', xaxis_title="Payment type", yaxis_title='%')
                    if datalabels == 'On':
                        fig3.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig3)
                if select_sector == "Support payment type":
                    df = df[df['Description2'] == select_sector]
                    fig = go.Figure()
                    for region in regions:
                        fig.add_trace(go.Bar(x=df['Description3'], y=df[region], name=region))
                    fig.update_layout(barmode='group', title='Support payment type', xaxis_title="Payment type", yaxis_title='%')
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig)
                if select_sector == "Aged under 25 / aged over 75":
                    df = df[df['Description2'] == select_sector]
                    dfpc = df[df['Unit'] == '%']
                    dfno = df[df['Unit'] == 'no.']
                    fig = go.Figure()
                    for region in regions:
                        fig.add_trace(go.Bar(x=dfpc['Description2'], y=dfpc[region], name=region))
                    fig.update_layout(barmode='group', title='Aged under 25 and aged over 75', xaxis_title="Age group", yaxis_title='%')
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig)
                    fig2 = go.Figure()
                    for region in regions:
                        fig2.add_trace(go.Bar(x=dfno['Description2'], y=dfno[region], name=region))
                    fig2.update_layout(barmode='group', title='Aged under 25 and aged over 75', xaxis_title="Age group", yaxis_title='no.')
                    if datalabels == 'On':
                        fig2.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig2)
                if select_sector == "Family type":
                    df = df[df['Description2'] == select_sector]
                    dfpc = df[df['Unit'] == '%']
                    dfno = df[df['Unit'] == 'no.']
                    fig = go.Figure()
                    for region in regions:
                        fig.add_trace(go.Bar(x=dfpc['Description3'], y=dfpc[region], name=region))
                    fig.update_layout(barmode='group', title='Family type', xaxis_title="Family type", yaxis_title='%')
                    if datalabels == 'On':
                        fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig)
                    fig2 = go.Figure()
                    for region in regions:
                        fig2.add_trace(go.Bar(x=dfno['Description3'], y=dfno[region], name=region))
                    fig2.update_layout(barmode='group', title='Family type', xaxis_title="Family type", yaxis_title='no.')
                    if datalabels == 'On':
                        fig2.update_traces(texttemplate='%{y:.2s}', textposition='inside')
                    st.plotly_chart(fig2)
        elif select_view == 'Housing affordability':
        #Description1 = Income units receiving CRA at 30 June
            df= df[df['Description1'] == 'Income units receiving CRA at 30 June']
            #sort year ascending
            df = df.sort_values(by=['Year'], ascending=True)
            #if Equity_Group null copy from Remoteness
            df['Equity_Group'] = df['Equity_Group'].fillna(df['Remoteness'])
            with col2:
                ha_filter = st.selectbox('Select group', df['Equity_Group'].unique(), index=4)
            df = df[df['Equity_Group'] == ha_filter]
            df1 = df[df['Description2'] == 'Paying more than 30% of income on rent']
            df2 = df[df['Description2'] == 'Paying more than 50% of income on rent']
            fig = go.Figure()
            for region in regions:
                fig.add_trace(go.Bar(x=[df1['Description4'], df['Year']], y=df1[region], name=region))
            fig.update_layout(barmode='group', title='Proportion recipients paying more than 30% of income on rent', yaxis_title='%')
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig)
            fig2 = go.Figure()
            for region in regions:
                fig2.add_trace(go.Bar(x=[df2['Description4'], df['Year']], y=df2[region], name=region))
            fig2.update_layout(barmode='group', title='Proportion recipients paying more than 50% of income on rent', yaxis_title='%')
            if datalabels == 'On':
                fig2.update_traces(texttemplate='%{y:.2s}', textposition='inside')
            st.plotly_chart(fig2)

    return
      
def ROGS_housing():
    st.markdown(f'Source: <a href="{ROGSHousingSourceURL}">{ROGSHousingSourceText}</a>', unsafe_allow_html=True)

    rogshousing = pd.read_csv(ROGSHousingdf, encoding='latin-1')

    rogshousing = rogshousing[rogshousing['Measure'] != 'Descriptive data']
    rogshousing = rogshousing[rogshousing['Measure'] != 'Survey response rates']
    rogshousing = rogshousing[rogshousing['Measure'] != 'Self-reported benefits of living in social housing - Public housing']
    rogshousing = rogshousing[rogshousing['Measure'] != 'Self-reported benefits of living in social housing - SOMIH']
    rogshousing = rogshousing[rogshousing['Measure'] != 'Self-reported benefits of living in social housing - Community housing']

    rogshousing = rogshousing.sort_values(by='Year', ascending=True)

    col1, col2 = st.columns(2)
    with col1: 
        measure = st.selectbox('Measure', rogshousing['Measure'].unique())
        filtered_data=rogshousing[rogshousing['Measure']==measure]
        if measure == 'Recurrent expenditure':
            filtered_data = filtered_data[filtered_data['Housing_Type'] != 'Community housing']
            filtered_data = filtered_data[filtered_data['Housing_Type'].notna()]
            filtered_data = filtered_data[filtered_data['Housing_Type'] != 'Indigenous community housing']

    with col1:
        housing_type = st.selectbox('Housing type', filtered_data['Housing_Type'].unique())
        filtered_data=filtered_data[filtered_data['Housing_Type']==housing_type]
    with col2:
        desc1 = st.selectbox('Description1', filtered_data['Description1'].unique())
        filtered_data=filtered_data[filtered_data['Description1']==desc1]
        desc2 = st.selectbox('Description2', filtered_data['Description2'].unique())
        filtered_data=filtered_data[filtered_data['Description2']==desc2]
    with col1:
        if len(filtered_data['Description3'].unique()) > 1:
            desc3 = st.selectbox('Description3', filtered_data['Description3'].unique())
            filtered_data=filtered_data[filtered_data['Description3']==desc3]
    with col2:
        if len(filtered_data['Description4'].unique()) > 1:
            desc4 = st.selectbox('Description4', filtered_data['Description4'].unique())
            filtered_data=filtered_data[filtered_data['Description4']==desc4]
    with col1:
        if len(filtered_data['Description5'].unique()) > 1:
            desc5 = st.selectbox('Description5', filtered_data['Description5'].unique())
            filtered_data=filtered_data[filtered_data['Description5']==desc5]
    with col2:
        if len(filtered_data['Description6'].unique()) > 1:
            desc6 = st.selectbox('Description6', filtered_data['Description6'].unique())
            filtered_data=filtered_data[filtered_data['Description6']==desc6]
    with col1:
        chart_type = st.radio('Chart type', ['Line chart', 'Bar chart'])
    with col2:
        st.markdown('<table style="background-color: yellow; font-weight: bold; font-style: italic"><tr><td>Series can be toggled on/off by clicking on the legend</td></tr></table>', unsafe_allow_html=True)

    if len(filtered_data['Total'].unique()) > 1:
        regions = ['Total', 'WA','NSW', 'Vic', 'Qld', 'WA', 'SA','Tas', 'ACT', 'NT']
    else:
        regions = ['Aust', 'WA', 'NSW', 'Vic', 'Qld', 'SA','Tas', 'ACT', 'NT']

    fig=go.Figure()

    if chart_type == 'Line chart':
        for region in regions:
            fig.add_trace(go.Scatter(x=filtered_data['Year'], y=filtered_data[region], name=region, mode='lines+markers'))
        fig.update_layout(title_text=f'{measure} - {desc1} {desc2}', yaxis=dict(title=filtered_data['Unit'].unique()[0]), xaxis=dict(title='Year'))

    else:
        for region in regions:
            fig.add_trace(go.Bar(x=filtered_data['Year'], y=filtered_data[region], name=region))
        fig.update_layout(title_text=f'{measure} - {desc1} {desc2}', yaxis=dict(title=filtered_data['Unit'].unique()[0]), xaxis=dict(title='Year'), barmode='group')
    st.plotly_chart(fig, use_container_width=True)
    
    return

def ROGS_homelessness():
    st.markdown(f'Source: <a href="{ROGSHomelessnessSourceURL}">{ROGSHomelessnessSourceText}</a>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    df = pd.read_csv(ROGSHomelessnessdf, encoding='latin-1')
    #sort year ascending
    df = df.sort_values(by='Year', ascending=True)
    df['Year'] = df['Year'].astype(str)

    df = df.rename(columns={'Aust': 'National'})
    regions = ['National', 'WA', 'Vic', 'Qld', 'SA', 'NSW', 'Tas', 'NT', 'ACT']
    #df long = melt on regions, value_name='Value', var_name='Region'
    cols = df.columns.tolist()
    #remove regions from cols
    for region in regions:
        cols.remove(region)
    dflong = pd.melt(df, id_vars=cols, value_vars=regions, var_name='Region', value_name='Value')

    #filter out measure = Rate of homeless people, Composition of support provided, Access of selected equity groups, Addressing client needs, Achievement of employment; education and/or training on exit, Achievement of income on exit,Clients at risk of homelessness who avoided homelessness,Support periods in which clients at risk of homelessness avoided homelessness,Achievement of independent housing on exit,Clients who return to homelessness after achieving housing, Clients who experience persistent homelessness
    df = df[df['Measure'] != 'Rate of homeless people']
    df = df[df['Measure'] != 'Composition of support provided']
    df = df[df['Measure'] != 'Access of selected equity groups']
    df = df[df['Measure'] != 'Addressing client needs']
    df = df[df['Measure'] != 'Achievement of employment; education and/or training on exit']
    df = df[df['Measure'] != 'Achievement of income on exit']
    df = df[df['Measure'] != 'Clients at risk of homelessness who avoided homelessness']
    df = df[df['Measure'] != 'Support periods in which clients at risk of homelessness avoided homelessness']
    df = df[df['Measure'] != 'Achievement of independent housing on exit']
    df = df[df['Measure'] != 'Clients who return to homelessness after achieving housing']
    df = df[df['Measure'] != 'Clients who experience persistent homelessness']


    with col1:
        select_measure = st.selectbox('Select measure', df['Measure'].unique())

    df = df[df['Measure'] == select_measure]
    df['Year'] = df['Year'].astype(str)

    if select_measure == "Recurrent expenditure":
        
        st.markdown('<table style="background-color: yellow; font-weight: bold; font-style: italic"><tr><td>Region series can be toggled on/off by clicking on the legend</td></tr></table>', unsafe_allow_html=True)
        ytitle = df['Unit'].unique()[0] + ' (' + df['Year_Dollars'].unique()[0] + ')'
        with col2:
            showas = st.radio('Show as', ['Total', 'Per person in population'], index=0, horizontal=True)
        if showas == 'Total':
            dfRE = df[df['Description2'] == 'Total recurrent real expenditure']
            charttitle = 'Total recurrent real expenditure'
        if showas == 'Per person in population':
            dfRE = df[df['Description2'] == 'Real expenditure per person in the residential population']
            charttitle = 'Real expenditure per person in the residential population'
        fig = go.Figure()
        for region in regions:
            fig.add_trace(go.Bar(x=dfRE['Year'], y=dfRE[region], name=region))
        fig.update_layout(barmode='group', title=charttitle, xaxis_title="Year", yaxis_title=ytitle)
        st.plotly_chart(fig)

    if select_measure == "Unmet need":
        filter1 = st.selectbox('Select filter', ['Accommodation services', 'Services other than accommodation'], index=0, key='filter1')
        
        st.markdown('<table style="background-color: yellow; font-weight: bold; font-style: italic"><tr><td>Region series can be toggled on/off by clicking on the legend</td></tr></table>', unsafe_allow_html=True)
        if filter1 == 'Accommodation services':
            df_fig2 = df[df['Description1'] == 'Average daily unassisted requests']
            df_fig2 = df_fig2.sort_values(by=['Year'], ascending=True)
            df_fig2 = df_fig2[df_fig2['Description2'] == 'Accommodation services']
            fig2 = go.Figure()
            for region in regions:
                fig2.add_trace(go.Bar(x=df_fig2['Year'], y=df_fig2[region], name=region))
            fig2.update_layout(barmode='group', title='Average daily unassisted requests', xaxis_title="Year", yaxis_title='Number')
            st.plotly_chart(fig2)
            df_fig1 = df[df['Description1'] == 'Accommodation services']
            #sort Year ascending
            df_fig1 = df_fig1.sort_values(by=['Year'], ascending=True)
            for Desc2 in df_fig1['Description2'].unique().tolist():
                fig1 = go.Figure()
                df_fig1_fil = df_fig1[df_fig1['Description2'] == Desc2]
                for region in regions:
                    fig1.add_trace(go.Bar(x=df_fig1_fil['Year'], y=df_fig1_fil[region], name=region))
                fig1.update_layout(barmode='group', title=Desc2, xaxis_title="Year", yaxis_title='Number')
                st.plotly_chart(fig1)
        if filter1 == 'Services other than accommodation':
            df_fig2 = df[df['Description1'] == 'Average daily unassisted requests']
            df_fig2 = df_fig2.sort_values(by=['Year'], ascending=True)
            df_fig2 = df_fig2[df_fig2['Description2'] == 'Services other than accommodation']
            fig2 = go.Figure()
            for region in regions:
                fig2.add_trace(go.Bar(x=df_fig2['Year'], y=df_fig2[region], name=region))
            fig2.update_layout(barmode='group', title='Average daily unassisted requests', xaxis_title="Year", yaxis_title='Number')
            st.plotly_chart(fig2)
            df_fig1 = df[df['Description1'] == 'Services other than accommodation']
            #sort Year ascending
            df_fig1 = df_fig1.sort_values(by=['Year'], ascending=True)
            for Desc2 in df_fig1['Description2'].unique().tolist():
                fig1 = go.Figure()
                df_fig1_fil = df_fig1[df_fig1['Description2'] == Desc2]
                for region in regions:
                    fig1.add_trace(go.Bar(x=df_fig1_fil['Year'], y=df_fig1_fil[region], name=region))
                fig1.update_layout(barmode='group', title=Desc2, xaxis_title="Year", yaxis_title='Number')
                st.plotly_chart(fig1)

    if select_measure == "Homelessness; by homelessness operational group":
        #fill Description2 null with persons
        df['Description2'] = df['Description2'].fillna('Persons')
        col1, col2, col3 = st.columns(3)
        with col1:
            Desc2 = st.selectbox('Select Description2 filter', df['Description2'].unique(), index=0)
            df = df[df['Description2'] == Desc2]
        with col2:
            if len(df['Description3'].unique()) > 1:
                Desc3 = st.selectbox('Select Description3 filter', df['Description3'].unique(), index=0)
                df = df[df['Description3'] == Desc3]
        with col3:
            if len(df['Description4'].unique()) > 1:
                Desc4 = st.selectbox('Select Description4 filter', df['Description4'].unique(), index=0)
                df = df[df['Description4'] == Desc4]

        #sort Year ascending
        df = df.sort_values(by=['Year'], ascending=True)
        #for region in regions, filter df for region, plotly bar, x=Year, y=Value, color=Region, group
        fig = go.Figure()
        yunits = df['Unit'].unique()[0]
        for region in regions:
            fig.add_trace(go.Bar(x=df['Year'], y=df[region], name=region))
        fig.update_layout(barmode='group', title='Homelessness; by homelessness operational group', xaxis_title="Year", yaxis_title=yunits)
        st.plotly_chart(fig)
    return

def external_resources():
    external = pd.read_excel('assets/External.xlsx', sheet_name='Sheet1')
    resource_filter = st.selectbox('Select resource group', external['Filter'].unique())
    external = external[external['Filter'] == resource_filter]
    for i in external.index:
        if external['Type'][i] == 'Link':
            
                st.markdown(f'<a href="{external["Reference link"][i]}">{external["Reference text"][i]}</a>', unsafe_allow_html=True)
        else:
            st.markdown(f'<h5>{external["caption"][i]}</h5>', unsafe_allow_html=True)
            try:
                file = 'assets/' + external['File'][i]
                st.image(file, use_column_width=True)
                st.markdown(f'<a href="{external["Reference link"][i]}">Source: {external["Reference text"][i]}</a>', unsafe_allow_html=True)
            except:
                pass
            
    return

def airbnb_wa():
    df_wa_total = pd.read_csv(AirbnbWATotaldf)
    df_wa_total['date'] = pd.to_datetime(df_wa_total['date'], format='%d/%m/%Y', dayfirst=True, errors='coerce')
    df_wa_total = df_wa_total.sort_values(by='date', ascending=True)
    df_wa_total = df_wa_total.rename(columns={'count_listings': 'count'})
    
    df_wa_total['date'] = df_wa_total['date'].astype(str)
    fig = go.Figure()
    for room_type in df_wa_total['room_type'].unique():
        df_room_type = df_wa_total[df_wa_total['room_type'] == room_type]
        fig.add_trace(go.Bar(x=df_room_type['date'].astype(str), y=df_room_type['count'], name=room_type))
    fig.update_layout(barmode='stack', xaxis={'categoryorder':'category ascending'})
    fig.update_layout(title='Number of Airbnb listings in WA by type', xaxis_title='', yaxis_title='Number of listings')
    fig.update_layout(margin=dict(b=0))
    st.plotly_chart(fig)
    fig2 = go.Figure()
    for room_type in df_wa_total['room_type'].unique():
        df_room_type = df_wa_total[df_wa_total['room_type'] == room_type]
        fig2.add_trace(go.Bar(x=df_room_type['date'].astype(str), y=df_room_type['price_median'], name=room_type))
    fig2.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig2.update_layout(title='Median price of Airbnb listings in WA', xaxis_title='', yaxis_title='Median price ($)')
    st.plotly_chart(fig2)
    fig3 = go.Figure()
    for room_type in df_wa_total['room_type'].unique():
        df_room_type = df_wa_total[df_wa_total['room_type'] == room_type]
        fig3.add_trace(go.Bar(x=df_room_type['date'].astype(str), y=df_room_type['price_mean'], name=room_type))
    fig3.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig3.update_layout(title='Mean price of Airbnb listings in WA', xaxis_title='', yaxis_title='Mean price ($)')
    st.plotly_chart(fig3)
    fig4 = go.Figure()
    for room_type in df_wa_total['room_type'].unique():
        df_room_type = df_wa_total[df_wa_total['room_type'] == room_type]
        fig4.add_trace(go.Bar(x=df_room_type['date'].astype(str), y=df_room_type['availability_365_median'], name=room_type))
    fig4.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig4.update_layout(title='Median availability of Airbnb listings in WA', xaxis_title='', yaxis_title='Median availability (days)')
    st.plotly_chart(fig4)
    fig5 = go.Figure()
    for room_type in df_wa_total['room_type'].unique():
        df_room_type = df_wa_total[df_wa_total['room_type'] == room_type]
        fig5.add_trace(go.Bar(x=df_room_type['date'].astype(str), y=df_room_type['availability_365_mean'], name=room_type))
    fig5.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig5.update_layout(title='Mean availability of Airbnb listings in WA', xaxis_title='', yaxis_title='Mean availability (days)')
    st.plotly_chart(fig5)
    return

def airbnb_geo():
    df_geo = pd.read_csv(AirbnbGeodf)
    df_geo = df_geo.rename(columns={'SA2_NAME_2016':'SA2', 'SA3_NAME_2016':'SA3', 'SA4_NAME_2016':'SA4', 'id_count': 'count'})
    st.markdown(f'#### Geographic filters')
    select_geo = st.radio('Select geography filter type:', ['Census areas (multi-level)', 'Federal electorate', 'LGA'], index=0)
    col1, col2, col3 = st.columns(3)
    if select_geo == 'Census areas (multi-level)':
        with col1:
            SA4 = st.multiselect('Select SA4', df_geo['SA4'].unique())
            if SA4:
                df_geo_fil = df_geo[df_geo['SA4'].isin(SA4)]
        with col2:
            if SA4:
                SA3 = st.multiselect('Select SA3', df_geo_fil['SA3'].unique(), default=df_geo_fil['SA3'].unique())
                df_geo_fil = df_geo_fil[df_geo_fil['SA3'].isin(SA3)]
            else:
                SA3 = st.multiselect('Select SA3', df_geo['SA3'].unique())
                if SA3:
                    df_geo_fil = df_geo[df_geo['SA3'].isin(SA3)]
                    if len(SA3) == len(df_geo['SA3'].unique()):
                        df_geo_fil = df_geo_fil.groupby(['date', 'room_type', 'SA4']).agg({'count': 'sum', 'price_mean': 'mean', 'availability_365_mean': 'mean', 'price_median': 'median', 'availability_365_median': 'median'}).reset_index()
        with col3:
            if SA3:
                SA2 = st.multiselect('Select SA2', df_geo_fil['SA2'].unique(), default=df_geo_fil['SA2'].unique())
                df_geo_fil = df_geo_fil[df_geo_fil['SA2'].isin(SA2)]
            else:
                SA2 = st.multiselect('Select SA2', df_geo['SA2'].unique())
                if SA2:
                    df_geo_fil = df_geo[df_geo['SA2'].isin(SA2)]
                    if len(SA2) == len(df_geo['SA2'].unique()):
                        df_geo_fil = df_geo_fil.groupby(['date', 'room_type', 'SA3']).agg({'count': 'sum', 'price_mean': 'mean', 'availability_365_mean': 'mean', 'price_median': 'median', 'availability_365_median': 'median'}).reset_index()
    elif select_geo == 'Federal electorate':
        fed_electorate = st.multiselect('Select federal electorate', df_geo['electorate'].unique())
        if fed_electorate:
            df_geo_fil = df_geo[df_geo['electorate'].isin(fed_electorate)]
            df_geo_fil = df_geo_fil.groupby(['date', 'room_type', 'electorate']).agg({'count': 'sum', 'price_mean': 'mean', 'availability_365_mean': 'mean', 'price_median': 'median', 'availability_365_median': 'median'}).reset_index()
    elif select_geo == 'LGA':
        LGA = st.multiselect('Select LGA', df_geo['lgaregion'].unique())
        if LGA:
            df_geo_fil = df_geo[df_geo['lgaregion'].isin(LGA)]
            df_geo_fil = df_geo_fil.groupby(['date', 'room_type', 'lgaregion']).agg({'count': 'sum', 'price_mean': 'mean', 'availability_365_mean': 'mean', 'price_median': 'median', 'availability_365_median': 'median'}).reset_index()
    try:
        room_type = st.multiselect('Select room type', df_geo_fil['room_type'].unique(), default=df_geo_fil['room_type'].unique())
        if room_type:
            df_geo_fil = df_geo_fil[df_geo_fil['room_type'].isin(room_type)]
    except:
        room_type = st.multiselect('Select room type', df_geo['room_type'].unique(), default=df_geo['room_type'].unique())
        if room_type:
            df_geo_fil = df_geo[df_geo['room_type'].isin(room_type)]
    fig = go.Figure()
    for room_type in df_geo_fil['room_type'].unique():
        df_room_type = df_geo_fil[df_geo_fil['room_type'] == room_type]
        fig.add_trace(go.Bar(x=df_geo_fil['date'], y=df_geo_fil['count'], name=room_type))
    fig.update_layout(barmode='stack', xaxis={'categoryorder':'category ascending'})
    fig.update_layout(title='Number of Airbnb listings in area by type', xaxis_title='', yaxis_title='Number of listings',margin=dict(b=0))
    st.markdown(f'*Hover values over bars in geographic filtered chart do not currently reflect single total for date, room type - currently showing multiple points for each suburb in area, to be corrected*')
    st.plotly_chart(fig)
    fig2 = go.Figure()
    for room_type in df_geo_fil['room_type'].unique():
        df_room_type = df_geo_fil[df_geo_fil['room_type'] == room_type]
        fig2.add_trace(go.Bar(x=df_room_type['date'], y=df_room_type['price_median'], name=room_type))
    fig2.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig2.update_layout(title='Median price of Airbnb listings in area', xaxis_title='', yaxis_title='Median price ($)')
    st.plotly_chart(fig2)
    fig3 = go.Figure()
    for room_type in df_geo_fil['room_type'].unique():
        df_room_type = df_geo_fil[df_geo_fil['room_type'] == room_type]
        fig3.add_trace(go.Bar(x=df_room_type['date'], y=df_room_type['price_mean'], name=room_type))
    fig3.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig3.update_layout(title='Mean price of Airbnb listings in area', xaxis_title='', yaxis_title='Mean price ($)')
    st.plotly_chart(fig3)
    fig4 = go.Figure()
    for room_type in df_geo_fil['room_type'].unique():
        df_room_type = df_geo_fil[df_geo_fil['room_type'] == room_type]
        fig4.add_trace(go.Bar(x=df_room_type['date'], y=df_room_type['availability_365_median'], name=room_type))
    fig4.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig4.update_layout(title='Median availability of Airbnb listings in area', xaxis_title='', yaxis_title='Median availability (days)')
    st.plotly_chart(fig4)
    fig5 = go.Figure()
    for room_type in df_geo_fil['room_type'].unique():
        df_room_type = df_geo_fil[df_geo_fil['room_type'] == room_type]
        fig5.add_trace(go.Bar(x=df_room_type['date'], y=df_room_type['availability_365_mean'], name=room_type))
    fig5.update_layout(barmode='group', xaxis={'categoryorder':'category ascending'})
    fig5.update_layout(title='Mean availability of Airbnb listings in area', xaxis_title='', yaxis_title='Mean availability (days)')
    st.plotly_chart(fig5)
    return

def delete_source_file(file):
    if os.path.exists(file):
        os.remove(file)
        return
    else:
        return

def update_log(latest_date, update_date, dataset):
    try:
        update_log = pd.read_excel(updatelogfile)
    except:
        update_log = pd.DataFrame(columns=['Dataset', 'Latest data point', 'Date last updated'])
    new_row = pd.DataFrame({'Dataset': [dataset], 'Latest data point': [latest_date], 'Date last updated': [update_date]})
    update_log = pd.concat([update_log, new_row], ignore_index=True)
    update_log['Latest data point'] = pd.to_datetime(update_log['Latest data point'], format='%d/%m/%Y')
    update_log['Date last updated'] = pd.to_datetime(update_log['Date last updated'], format='%d/%m/%Y')
    update_log = update_log.sort_values(by=['Latest data point', 'Date last updated'], ascending=False).drop_duplicates(subset=['Dataset'], keep='first')
    update_log['Latest data point'] = update_log['Latest data point'].dt.strftime('%d/%m/%Y')
    update_log['Date last updated'] = update_log['Date last updated'].dt.strftime('%d/%m/%Y')                            
    update_log.to_excel(updatelogfile, index=False)
    book = openpyxl.load_workbook(updatelogfile)
    sheet = book.active
    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length
    book.save(updatelogfile)
    book.close()
    return

def upload_data():
    col1, col2, col3 = st.columns(3)
    with col1:
        select_data_to_upload = st.selectbox('Select the data/content to upload', ['ROGS', 'Monthly SHS', 'Airbnb', 'Waitlist - WA total', 'Waitlist - breakdowns', 'Images or links'])
    if select_data_to_upload == 'Monthly SHS':
        with col2:
            st.markdown(f'**Step 1:**<a href="https://www.aihw.gov.au/reports/homelessness-services/specialist-homelessness-services-monthly-data/data">  Download **Data Tables** from Government site to your computer - follow this link, click DOWNLOAD button </a>', unsafe_allow_html=True)
        with col3:
            SHSnew = st.file_uploader("Step 2 - upload SHS file")
        if SHSnew is not None:
            source_file = pd.ExcelFile(SHSnew)
            import_shs_data(source_file)
    if select_data_to_upload == 'ROGS':
        st.markdown(f'**Step 1:**')
        st.markdown(f'Download **CSV** files requiring update, under Part G at <a href="https://www.pc.gov.au/ongoing/report-on-government-services">this website</a>', unsafe_allow_html=True)
        st.markdown(f'**Step 2:**')
        st.markdown(f'Upload files below, ensuring you select correct file for Sector Overview, Housing and Homelessness. If any do not require update, you do not need to upload them.', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            ROGSsectornew = st.file_uploader("Select Sector Overview file")
        with col2:
            ROGShousingnew = st.file_uploader("Select Housing file")
        with col3:
            ROGSHomelessnessnew = st.file_uploader("Select Homelessness file")
        if ROGSsectornew is not None:
            ROGSsector = pd.read_csv(ROGSsectornew, encoding='latin-1')
            ROGSsector.to_csv(ROGSSectordf, index=False)
            latest_date = ROGSsector['Year'].max()
            latest_date = latest_date[-2:]
            latest_date = f'30/06/20{latest_date}'
            latest_date = pd.to_datetime(latest_date, format='%d/%m/%Y', dayfirst=True)
            update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
            update_log(latest_date, update_date, 'ROGS Sector Overview')
        if ROGShousingnew is not None:
            ROGShousing = pd.read_csv(ROGShousingnew, encoding='latin-1')
            ROGShousing.to_csv(ROGSHousingdf, index=False)
            latest_date = ROGShousing['Year'].max()
            latest_date = latest_date[-2:]
            latest_date = f'30/06/20{latest_date}'
            latest_date = pd.to_datetime(latest_date, format='%d/%m/%Y', dayfirst=True)
            update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
            update_log(latest_date, update_date, 'ROGS Housing')
        if ROGSHomelessnessnew is not None:
            ROGSHomelessness = pd.read_csv(ROGSHomelessnessnew, encoding='latin-1')
            ROGSHomelessness.to_csv(ROGSHomelessnessdf, index=False)
            latest_date = ROGSHomelessness['Year'].max()
            latest_date = latest_date[-2:]
            latest_date = f'30/06/20{latest_date}'
            latest_date = pd.to_datetime(latest_date, format='%d/%m/%Y', dayfirst=True)
            update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
            update_log(latest_date, update_date, 'ROGS Homelessness')
    if select_data_to_upload == 'Airbnb':
        st.markdown(f'**Step 1:**')
        st.markdown(f'Download **listings.csv** under **Western Australia** files from <a href="http://insideairbnb.com/get-the-data.html">this website</a> - Western Australia is towards the bottom of the long page; press **Ctrl** + **F** keys and type in "Western Au" to jump there', unsafe_allow_html=True)
        st.markdown(f'**Step 2:**')
        filename = st.date_input('Enter date of data from Inside Airbnb website, which is listed above the download links', format="YYYY-MM-DD")
        filename = filename.strftime('%Y-%m-%d')
        st.markdown(f'**Step 3:**')
        airbnbnew = st.file_uploader("Select downloaded file")
        if airbnbnew is not None:
            airbnb = pd.read_csv(airbnbnew)
            get_airbnb(airbnb, filename)
            state_total(airbnb, filename)
    if select_data_to_upload == 'Waitlist - WA total':
        st.markdown(f'**Step 1:**')
        st.markdown(f'Waitlist numbers can be found in Housing portfolio Parliamentary Questions - search for waitlist and answered questions in Housing portfolio <a href="{waitlist_sourceURL}"> here</a>', unsafe_allow_html=True)
        st.markdown(f'**Step 2:**')
        st.markdown(f'Enter numbers and date below - press **Add data** button to add to dataset - data will be previewed and you will need to confirm', unsafe_allow_html=True)
        st.markdown(f'*Note - use **Tab** key to move between entry fields for quick filling*', unsafe_allow_html=True)
        df = pd.read_csv(newWaitlistData)
        col1, col2, col3, col4, col5, col6 = st.columns(6)
        with col1:
            newDate = st.date_input('Enter date of data', format="DD/MM/YYYY")
            
        with col2:
            newTotalApp = st.number_input('Total applications', min_value=0, value=0)
        with col3:    
            newPriApp = st.number_input('Priority applications', min_value=0, value=0)
        with col4:
            newTotalInd = st.number_input('Total individuals', min_value=0, value=0)
        with col5:
            newPriInd = st.number_input('Priority individuals', min_value=0, value=0)
        with col1:
            add_data = st.button('Add data')
        if add_data:
            newRow = pd.DataFrame({'Date': [newDate], 'Total applications': [newTotalApp], 'Priority applications': [newPriApp], 'Total individuals': [newTotalInd], 'Priority individuals': [newPriInd]})
            df = pd.concat([df, newRow], ignore_index=True)
            st.markdown(f'Preview')
            st.warning('Confirm data to be added')
            st.write(df)
            col1, col2, col3, col4= st.columns(4)
            with col1:
                confirm = st.button('Confirm')
            with col2:
                cancel = st.button('Cancel')
            if confirm:
                df.to_csv(newWaitlistData, index=False)
                import_waitlist_data()
                st.write('Data added')
            if cancel:
                #delete rows from newWaitlistData
                df = pd.read_csv(newWaitlistData)
                df = df[:-1]
        st.markdown(f'</br>', unsafe_allow_html=True)
        col1, col2= st.columns(2)
        with col1:
            st.markdown(f'***CSV file for in-Excel entry option, or override of incorrect updates***')
            with open(SimpleWaitlistData, 'rb') as file:
                btn = st.download_button(
                label="Download Waitlist Trend Data",
                data=file,
                file_name="Waitlist_trend.csv",
                mime="text/csv"
                )

        with col1:
            manualdata = st.file_uploader("After making changes and saving (as CSV), upload file here to update dataset")
        if manualdata is not None:
            manualdata = pd.read_csv(manualdata)
            manualdata.to_csv(SimpleWaitlistData, index=False)
            import_waitlist_manual(manualdata)
            st.write('Data updated')
    if select_data_to_upload == 'Waitlist - breakdowns':
        df = pd.read_csv(Waitlist_breakdownsdf)
        items = df['Item'].unique().tolist()
        #remove 'Waiting time (WA total only -extra data point)'
        items.remove('Waiting time (WA total only -extra data point)')
        items.remove('New tenancies')

        with col2: 
            select_item = st.selectbox('Select item', items)
        if select_item == "Dwelling need":
            template = 'DATA/SOURCE DATA/TEMPLATES/DwellingNeedTemplate.xlsx'
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                with open(template, 'rb') as file:
                    btn = st.download_button(
                        label="Download Dwelling Need Template",
                        data=file,
                        file_name="DwellingNeedTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col1:
                file = st.file_uploader("Upload completed template")
            with col2:
                date_data = st.date_input('Enter date of data', format="DD/MM/YYYY", key='date_data')
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                st.markdown(f'</br>', unsafe_allow_html=True)
                add_dwelling_need_data = st.button('Add data')
            if add_dwelling_need_data:
                df = pd.read_excel(file)
                df0 = pd.read_csv(Waitlist_breakdownsdf)
                df['Item'] = "Dwelling need"
                df['Date'] = date_data
                df = df.rename(columns={'Total applications': 'Total Waitlist - Applications', 'Total individuals': 'Total Waitlist - Individuals', 'Priority applications': 'Priority Waitlist - Applications', 'Priority individuals': 'Priority Waitlist - Individuals', 'Dwelling need': 'Detail'})
                df = df.melt(id_vars=['Detail', 'Date', 'Item'], var_name='Category0', value_name='Value')
                df[['Category1', 'Category2']] = df['Category0'].str.split(' - ', expand=True)
                df = df.drop(columns=['Category0'])
                df = df.rename(columns={'Category1': 'Category'})
                df = df.rename(columns={'Category2': 'Subcategory'})
                df['Region'] = 'WA'
                df = pd.concat([df0, df], ignore_index=True)
                df.to_csv(Waitlist_breakdownsdf, index=False)
                st.write('Data added')
        if select_item == "Region need":
            template = 'DATA/SOURCE DATA/TEMPLATES/RegionNeedTemplate.xlsx'
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                with open(template, 'rb') as file:
                    btn = st.download_button(
                        label="Download Region Need Template",
                        data=file,
                        file_name="RegionNeedTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col1:
                file = st.file_uploader("Upload completed template")
            with col2:
                date_data = st.date_input('Enter date of data', format="DD/MM/YYYY", key='date_data')
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                st.markdown(f'</br>', unsafe_allow_html=True)
                add_region_need_data = st.button('Add data')
            if add_region_need_data:
                df = pd.read_excel(file)
                df0 = pd.read_csv(Waitlist_breakdownsdf)
                df['Item'] = "Region need"
                df = df.rename(columns={'Region need': 'Region'})
                df['Date'] = date_data
                df = df.melt(id_vars=['Region', 'Date', 'Item'], var_name='Category0', value_name='Value')
                df[['Category', 'Subcategory']] = df['Category0'].str.split(' - ', expand=True)
                df = df.drop(columns=['Category0'])
                df = pd.concat([df0, df], ignore_index=True)
                df.to_csv(Waitlist_breakdownsdf, index=False)
                st.write('Data added')
        if select_item == "Waiting time by dwelling need":
            template = 'DATA/SOURCE DATA/TEMPLATES/WaitingTimebyDwellingNeedTemplate.xlsx'
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                with open(template, 'rb') as file:
                    btn = st.download_button(
                        label="Download Waiting Time by Dwelling Need Template",
                        data=file,
                        file_name="WaitingTimebyDwellingNeedTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col1:
                file = st.file_uploader("Upload completed template")
            with col2:
                date_data = st.date_input('Enter date of data', format="DD/MM/YYYY", key='date_data')
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                st.markdown(f'</br>', unsafe_allow_html=True)
                add_waiting_time_by_dwelling_need_data = st.button('Add data')
            if add_waiting_time_by_dwelling_need_data:
                df = pd.read_excel(file)
                df0 = pd.read_csv(Waitlist_breakdownsdf)
                df['Item'] = "Waiting time by dwelling need"
                df = df.rename(columns={'Waiting time by dwelling need': 'Detail'})
                df['Date'] = date_data
                df = df.melt(id_vars=['Detail', 'Date', 'Item'], var_name='Category0', value_name='Value')
                df[['Category', 'Subcategory']] = df['Category0'].str.split(' - ', expand=True)
                df = df.drop(columns=['Category0'])
                df = pd.concat([df0, df], ignore_index=True)
                df.to_csv(Waitlist_breakdownsdf, index=False)
                st.write('Data added')
        if select_item == "Waiting time by region":
            template = 'DATA/SOURCE DATA/TEMPLATES/WaitingTimebyRegionTemplate.xlsx'
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                with open(template, 'rb') as file:
                    btn = st.download_button(
                        label="Download Waiting Time by Region Template",
                        data=file,
                        file_name="WaitingTimebyRegionTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col1:
                file = st.file_uploader("Upload completed template")
            with col2:
                date_data = st.date_input('Enter date of data', format="DD/MM/YYYY", key='date_data')
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                st.markdown(f'</br>', unsafe_allow_html=True)
                add_waiting_time_by_region_data = st.button('Add data')
            if add_waiting_time_by_region_data:
                df = pd.read_excel(file)
                df0 = pd.read_csv(Waitlist_breakdownsdf)
                df['Item'] = "Waiting time by region"
                df = df.rename(columns={'Waiting time by region': 'Region'})
                df['Date'] = date_data
                df = df.melt(id_vars=['Region', 'Date', 'Item'], var_name='Category0', value_name='Value')
                df[['Category', 'Subcategory']] = df['Category0'].str.split(' - ', expand=True)
                df = df.drop(columns=['Category0'])
                df = pd.concat([df0, df], ignore_index=True)
                df.to_csv(Waitlist_breakdownsdf, index=False)
                st.write('Data added')
        if select_item == "New tenancies by dwelling need":
            template = 'DATA/SOURCE DATA/TEMPLATES/NewTenanciesbyDwellingNeedTemplate.xlsx'
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                with open(template, 'rb') as file:
                    btn = st.download_button(
                        label="Download New Tenancies by Dwelling Need Template",
                        data=file,
                        file_name="NewTenanciesbyDwellingNeedTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col1:
                file = st.file_uploader("Upload completed template")
            with col2:
                date_data = st.date_input('Enter date of data', format="DD/MM/YYYY", key='date_data')
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                st.markdown(f'</br>', unsafe_allow_html=True)
                add_new_tenancies_by_dwelling_need_data = st.button('Add data')
            if add_new_tenancies_by_dwelling_need_data:
                df = pd.read_excel(file)
                df0 = pd.read_csv(Waitlist_breakdownsdf)
                df['Item'] = "New tenancies by dwelling need"
                df = df.rename(columns={'New tenancies by dwelling need': 'Detail'})
                df['Date'] = date_data
                df = df.melt(id_vars=['Detail', 'Date', 'Item'], var_name='Category0', value_name='Value')
                df[['Category', 'Subcategory']] = df['Category0'].str.split(' - ', expand=True)
                df = df.drop(columns=['Category0'])
                df = pd.concat([df0, df], ignore_index=True)
                df.to_csv(Waitlist_breakdownsdf, index=False)
                st.write('Data added')
        if select_item == "New tenancies by region":
            template = 'DATA/SOURCE DATA/TEMPLATES/NewTenanciesbyRegionTemplate.xlsx'
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                with open(template, 'rb') as file:
                    btn = st.download_button(
                        label="Download New Tenancies by Region Template",
                        data=file,
                        file_name="NewTenanciesbyRegionTemplate.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            with col1:
                file = st.file_uploader("Upload completed template")
            with col2:
                date_data = st.date_input('Enter date of data', format="DD/MM/YYYY", key='date_data')
            with col3:
                st.markdown(f'</br>', unsafe_allow_html=True)
                st.markdown(f'</br>', unsafe_allow_html=True)
                add_new_tenancies_by_region_data = st.button('Add data')
            if add_new_tenancies_by_region_data:
                df = pd.read_excel(file)
                df0 = pd.read_csv(Waitlist_breakdownsdf)
                df['Item'] = "New tenancies by region"
                df = df.rename(columns={'New tenancies by region': 'Region'})
                df['Date'] = date_data
                df = df.melt(id_vars=['Region', 'Date', 'Item'], var_name='Category0', value_name='Value')
                df[['Category', 'Subcategory']] = df['Category0'].str.split(' - ', expand=True)
                df = df.drop(columns=['Category0'])
                df = pd.concat([df0, df], ignore_index=True)
                df.to_csv(Waitlist_breakdownsdf, index=False)
                st.write('Data added')
    if select_data_to_upload == 'Images or links':
        reffile = pd.read_excel('assets/External.xlsx')
        col1, col2 = st.columns(2)
        with col1:
            type = st.radio('Select type of content to upload', ['Image', 'Link'], horizontal=True)
        if type == 'Image':
            filterimage = reffile[reffile['Type'] == 'Image']
            filterimagelist = filterimage['Filter'].unique()
            filterimagelist = filterimagelist.tolist()
            #append other
            filterimagelist.append('Other')
            with col2:    
                selectfilter = st.selectbox('Select filter', filterimagelist)
            if selectfilter == 'Other':
                selectfilter = st.text_input('Enter filter')
            col1, col2 = st.columns(2)
            with col1:
                image = st.file_uploader("Select image file")
                ref_text = st.text_input('Enter reference text')
            with col2:
                image_name = st.text_input('Enter image filename')
                caption = st.text_input('Enter heading for image')
                ref_link = st.text_input('Enter reference link')
            
            if image is not None:
                upload_image = st.button('Upload image')
                if upload_image:
                    saveimageas = f'DATA/IMAGES/{image_name}.png'
                    image = PIL.Image.open(image)
                    image.save(saveimageas)
                    new_row = pd.DataFrame({'Filter': [selectfilter], 'caption': [caption], 'File': [saveimageas], 'Reference text': [ref_text], 'Reference link': [ref_link]})
                    reffile = pd.concat([reffile, new_row], ignore_index=True)
                    reffile.to_excel('assets/External.xlsx', index=False)
                    st.write('Image uploaded')
        if type == 'Link':
            filterlink = reffile[reffile['Type'] == 'Link']
            filterlinklist = filterlink['Filter'].unique()
            filterlinklist = filterlinklist.tolist()
            #append other
            filterlinklist.append('Other')
            with col2:
                selectfilter = st.selectbox('Select filter', filterlinklist)
            if selectfilter == 'Other':
                selectfilter = st.text_input('Enter filter')
            col1, col2 = st.columns(2)
            with col1:
                ref_text = st.text_input('Enter link text')
                ref_link = st.text_input('Enter link URL')
            upload_link = st.button('Add link')
            if upload_link:
                new_row = pd.DataFrame({'Filter': [selectfilter], 'Reference text': [ref_text], 'Reference link': [ref_link]})
                reffile = pd.concat([reffile, new_row], ignore_index=True)
                st.write('Link added')
                reffile.to_excel('assets/External.xlsx', index=False)
        st.markdown(f'</br>', unsafe_allow_html=True)
        st.markdown(f'***Reorder or remove external content***')
        st.markdown(f'To delete or re-order, download external reference file below, save changes and re-upload the file. Content is displayed in the order it appears in the file.')
        col1, col2 = st.columns(2)
        with col1:
            with open('assets/External.xlsx', 'rb') as file:
                    btn = st.download_button(
                    label="Download external content reference file",
                    data=file,
                    file_name="External.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    ) 
            file = st.file_uploader("Upload replacement reference file")
        if file is not None:
            reffile = pd.read_excel(file)
            reffile.to_excel('assets/External.xlsx', index=False)
            st.write('File uploaded')
        
    return

def as_text(value):
    if value is None:
        return ""
    return str(value)

def quarter_to_date(quarter):
    year, q = quarter.split('-')
    if q == 'Q1':
        return f'31/03/{year}'
    elif q == 'Q2':
        return f'30/06/{year}'
    elif q == 'Q3':
        return f'30/09/{year}'
    elif q == 'Q4':
        return f'31/12/{year}'

def group_age(age_group):
    if age_group.endswith('+'):
        lower_age_limit = int(age_group[:-1])
    elif age_group == 'All ages':
        return age_group
    else:
        lower_age_limit = int(age_group.split('-')[0])
    if lower_age_limit >= 65:
        return '65+'
    else:
        return age_group

def new_pop_file(file):
    Population_State_Sex_Age = pd.read_csv(file)
    Population_State_Sex_Age = Population_State_Sex_Age.rename(columns={'SEX: Sex': 'Sex', 'AGE: Age': 'Age group', 'TIME_PERIOD: Time Period': 'Quarter', 'REGION: Region': 'Region', 'OBS_VALUE': 'Population'})
    Population_State_Sex_Age = Population_State_Sex_Age.drop(columns=['DATAFLOW', 'MEASURE: Measure', 'FREQ: Frequency', 'UNIT_MEASURE: Unit of Measure', 'OBS_STATUS: Observation Status', 'OBS_COMMENT: Observation Comment'])
    Population_State_Sex_Age['Date'] = Population_State_Sex_Age['Quarter'].apply(quarter_to_date)
    Population_State_Sex_Age = Population_State_Sex_Age.drop(columns=['Quarter'])
    Population_State_Sex_Age['Sex'] = Population_State_Sex_Age['Sex'].map({
    '1: Males': 'Male',
    '2: Females': 'Female',
    '3: Persons': 'Total'
    })
    Population_State_Sex_Age['Age group'] = Population_State_Sex_Age['Age group'].str.split(': ').str[1]
    Population_State_Sex_Age['Region'] = Population_State_Sex_Age['Region'].map({
    '1: New South Wales': 'NSW',
    '2: Victoria': 'Vic',
    '3: Queensland': 'Qld',
    '4: South Australia': 'SA',
    '5: Western Australia': 'WA',
    '6: Tasmania': 'Tas',
    '7: Northern Territory': 'NT',
    '8: Australian Capital Territory': 'ACT',
    'AUS: Australia': 'National'
    })
    Population_State_Sex_Age['Date'] = pd.to_datetime(Population_State_Sex_Age['Date'], format='%d/%m/%Y')
    Population_State_Sex_Age = Population_State_Sex_Age.sort_values(by='Date', ascending=True)
    pivot_df = Population_State_Sex_Age.pivot_table(
        index=['Date', 'Sex', 'Age group'], 
        columns='Region', 
        values='Population',
        fill_value=0
    ).reset_index()

    pivot_df.columns = [f'{col}_Population' if col in ['NSW', 'Vic', 'Qld', 'WA', 'SA', 'Tas', 'ACT', 'NT', 'National'] else col for col in pivot_df.columns]

    Population_State_Sex_Age = pivot_df.rename(columns={'NSW': 'NSW_Population', 'Vic': 'Vic_Population', 'Qld': 'Qld_Population', 'WA': 'WA_Population', 'SA': 'SA_Population', 'Tas': 'Tas_Population', 'ACT': 'ACT_Population', 'NT': 'NT_Population', 'National': 'National_Population'})

    Population_State_Sex_Age['Age group'] = Population_State_Sex_Age['Age group'].apply(group_age)

    Population_State_Sex_Age = Population_State_Sex_Age.groupby(['Age group', 'Sex', 'Date']).agg({
        'NSW_Population': 'sum',
        'Vic_Population': 'sum',
        'Qld_Population': 'sum',
        'WA_Population': 'sum',
        'SA_Population': 'sum',
        'Tas_Population': 'sum',
        'ACT_Population': 'sum',
        'NT_Population': 'sum',
        'National_Population': 'sum'
    }).reset_index()

    latest_date = Population_State_Sex_Age['Date'].max()
    latest_date = pd.to_datetime(latest_date)
    try:
        current_file = pd.read_csv('DATA/PROCESSED DATA/Population/Population_State_Sex_Age_to_65+.csv')
    except:
        current_file = Population_State_Sex_Age
    
    latest_current_date = current_file['Date'].max()
    latest_current_date = pd.to_datetime(latest_current_date)

    if latest_date < latest_current_date:
        return
    else:
        Population_State_Sex_Age.to_csv(PopulationStateSexAge65df, index=False)
        latest_date = latest_date.strftime('%d/%m/%Y')
        update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
        update_log(latest_date, update_date, 'Population (by State, Sex, Age to 65+)')
    delete_source_file(PopulationNewFile)
    total(Population_State_Sex_Age)
    monthlyStatetotal()
    return

def total(df):
    df = df[df['Age group'] == 'All ages']
    df = df.drop(columns='Age group')
    save_to = 'DATA/PROCESSED DATA/Population/Population_State_Sex_Total'
    df.to_csv(save_to + '.csv', index=False)
    df = df[df['Sex'] == 'Total']
    df = df.drop(columns='Sex')
    save_to = 'DATA/PROCESSED DATA/Population/Population_State_Total'
    df.to_csv(save_to + '.csv', index=False)
    monthlyStatetotal(save_to)
    columns = df.columns.tolist()
    columns.remove('WA_Population')
    columns.remove('Date')
    df = df.drop(columns=columns)
    df = df.rename(columns={'WA_Population': 'Population'})
    save_to = 'DATA/PROCESSED DATA/Population/Population_WA_Total'
    df.to_csv(save_to + '.csv', index=False)
    return

def import_population_data():
    try:
        new_pop_file(PopulationNewFile)
    except:
        pass
    return

def monthlyStatetotal():
    df = pd.read_csv('DATA/PROCESSED DATA/Population/Population_State_Total.csv')
    df['Date'] = pd.to_datetime(df['Date'], format='%Y-%m-%d')
    df = df.sort_values(by='Date', ascending=True)
    df = df.set_index('Date').resample('M').mean().interpolate(method='linear').reset_index()
    df['Date'] = df['Date'].dt.strftime('%d/%m/%Y')
    df = df.round(0)
    df.to_csv('DATA/PROCESSED DATA/Population/Population_State_Total_monthly.csv', index=False)
    return

def get_SHS(source_file):
    xls = pd.ExcelFile(source_file)
    all_sheets = {sheet_name: pd.read_excel(xls, sheet_name, header=3) for sheet_name in xls.sheet_names}
    xls.close()
    for sheet_name, sheet in all_sheets.items():
        if len(sheet) > 100:
            sheet = sheet.drop(sheet.index[-2:])
            for col in sheet.columns:
                if sheet[col].dtype == 'object':
                    sheet[col] = sheet[col].str.replace(chr(8211), "-").str.replace(chr(8212), "-")
            save_sheet_name = sheet_name.replace(' ', '_')
            sheet.to_csv('DATA/PROCESSED DATA/SHS/SHS_' + save_sheet_name + '.csv', index=False)
            all_sheets.update({sheet_name: sheet})

def find_csv_filenames(path_to_dir, prefix, suffix):
    filenames = os.listdir(path_to_dir)
    return [ filename for filename in filenames if filename.endswith( suffix ) and filename.startswith(prefix) ]

def convert_case(df):
    # Convert column names to uppercase
    df.columns = [col.upper() for col in df.columns]
    
    # Convert string values in all columns to uppercase
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].str.capitalize()

    return df

def identify_ignore_columns(dataframes_dict):
    ignore_columns = set()
    for _, df in dataframes_dict.items():
        for column in df.columns:
            if df[column].dtype in ['int64', 'float64']:
                ignore_columns.add(column)
            elif 'datetime64' in str(df[column].dtype):
                ignore_columns.add(column)
            elif column == 'MONTH':  # specifically ignore 'MONTH' column
                ignore_columns.add(column)
    return list(ignore_columns)

def load_and_preprocess_data(path_to_dir, prefix, suffix):
    filenames = find_csv_filenames(path_to_dir, prefix, suffix)
    processed_dataframes = {}

    for filename in filenames:
        df_name = filename.replace('.csv', '')
        df = pd.read_csv(path_to_dir + '/'+ filename)
        df = convert_case(df)

        cols_to_check = ['NSW','VIC','QLD','WA','SA','TAS','ACT','NT', 'NATIONAL']
        df = df.dropna(subset=cols_to_check)

        ignore_cols = identify_ignore_columns({df_name: df})

        check_for_nan_cols = [col for col in df.columns if col not in ignore_cols]

        df = df.dropna(subset=check_for_nan_cols)

        if 'AGE GROUP' in df.columns:
            df['AGE GROUP'] = df['AGE GROUP'].str.replace(chr(45), "-").str.replace(chr(8211), "-")
            df['AGE GROUP'] = df['AGE GROUP'].astype(str)
            if 'All females' in df['AGE GROUP'].unique() or 'All males' in df['AGE GROUP'].unique():
                df = df[~df['AGE GROUP'].isin(['All females', 'All males'])]
                df['AGE GROUP'] = df['AGE GROUP'].str.replace(" years", "")
                df.loc[df['AGE GROUP'] == '15-17', 'AGE GROUP'] = '15-19'
                df.loc[df['AGE GROUP'] == '18-19', 'AGE GROUP'] = '15-19'
                object_cols = [col for col in df.columns if df[col].dtype == 'object']
                datetime_cols = [col for col in df.columns if 'datetime64' in str(df[col].dtype)]
                numeric_cols = [col for col in df.columns if df[col].dtype in ['int64', 'float64']]
                df = df.groupby(object_cols + datetime_cols)[numeric_cols].sum().reset_index()

        if 'MONTH' in df.columns:
            df['DATE'] = '20' + df['MONTH'].str[3:5] + '-' + df['MONTH'].str[0:3] + '-01'
            df['DATE'] = pd.to_datetime(df['DATE'], format='%Y-%b-%d')
            df['DATE'] = df['DATE'] + pd.offsets.MonthEnd(0)

        df = df.sort_values(by='DATE', ascending=True)
        df['DATE'] = df['DATE'].dt.strftime('%d/%m/%Y')

        processed_dataframes[df_name] = df
    return processed_dataframes

def merge_and_calculate(processed_dataframes, Population_Sex_Age, Population_Sex, Population_Total):

    pop_dfs = {}
    pop_dfs['Sex_Age'] = Population_Sex_Age
    pop_dfs['Sex'] = Population_Sex
    pop_dfs['Total'] = Population_Total

    for pop_df_name, pop_df in pop_dfs.items():
        pop_df.columns = [col.upper() for col in pop_df.columns]
        pop_df['DATE'] = pd.to_datetime(pop_df['DATE'], format='%Y-%m-%d', errors='coerce')
        pop_df = pop_df.set_index('DATE')
        pop_df_name = pop_df

    Population_Sex_Age['AGE GROUP'] = Population_Sex_Age['AGE GROUP'].str.replace(chr(45), "-").str.replace(chr(8211), "-")

    regions = ['NSW', 'VIC', 'QLD', 'WA', 'SA', 'TAS', 'ACT', 'NT']
    SHS_with_population_calcs = {}

    for df_name, df in processed_dataframes.items():
        df['DATE'] = pd.to_datetime(df['DATE'], format='%d/%m/%Y', dayfirst=True)
        if 'AGE GROUP' in df.columns:
            df['JoinLeft'] = df['DATE'].astype(str) + ' ' + df['SEX'].astype(str) + ' ' + df['AGE GROUP'].astype(str)
            Population_Sex_Age['JoinRight'] = Population_Sex_Age['DATE'].astype(str) + ' ' + Population_Sex_Age['SEX'].astype(str) + ' ' + Population_Sex_Age['AGE GROUP'].astype(str)
            merged_df = pd.merge(df, Population_Sex_Age, left_on=['JoinLeft'], right_on=['JoinRight'], how='left')
            merged_df = merged_df.sort_values(by=['SEX_y', 'AGE GROUP_y', 'DATE_y'])
            
        else:
            if 'SEX' in df.columns:
                df['JoinLeft'] = df['DATE'].astype(str) + ' ' + df['SEX'].astype(str)
                Population_Sex['JoinRight'] = Population_Sex['DATE'].astype(str) + ' ' + Population_Sex['SEX'].astype(str)
                merged_df = pd.merge(df, Population_Sex, left_on=['JoinLeft'], right_on=['JoinRight'], how='left')
                merged_df = merged_df.sort_values(by=['SEX_y', 'DATE_y']) 
            else:
                merged_df = pd.merge(df, Population_Total, left_on=['DATE'], right_on=['DATE'], how='left')
                merged_df = merged_df.sort_values(by=['DATE_y'])
            
        pop_cols = [col for col in merged_df.columns if col.endswith('_POPULATION')]
        merged_df[pop_cols] = merged_df[pop_cols].ffill(axis=1)
        merged_df = merged_df.sort_values(by=['DATE_x'])
        merged_df = merged_df.fillna(method='ffill')
        merged_df = merged_df.drop(columns=['JoinLeft', 'JoinRight'])
        merged_df = merged_df.loc[:,~merged_df.columns.str.endswith('_y')]
        merged_df = merged_df.rename(columns=lambda x: x.replace('_x', '') if x.endswith('_x') else x)
        cols = list(merged_df.columns)
        cols.insert(0, cols.pop(cols.index('DATE')))
        merged_df = merged_df[cols]

        merged_df['NATIONAL_PER_10k'] = merged_df['NATIONAL'] / merged_df['NATIONAL_POPULATION'] * 10000
        for region in regions:
            population_column_name = f"{region}_POPULATION"
            per_10000_column = f"{region}_PER_10k"
            merged_df[per_10000_column] = merged_df[region] / merged_df[population_column_name] * 10000
            proportion_of_national_column = f"{region}_PROPORTION_OF_NATIONAL"
            merged_df[proportion_of_national_column] = (merged_df[region] / merged_df['NATIONAL']) * 100
            proportion_of_national_per_10000_column = f"{region}_PROPORTION_OF_NATIONAL_PER_10k"
            merged_df[proportion_of_national_per_10000_column] = (merged_df[per_10000_column] / merged_df['NATIONAL_PER_10k']) * 100
            prop_national_pop_column = f"{region}_PROPORTION_OF_NATIONAL_POPULATION"     
            merged_df[prop_national_pop_column] = (merged_df[population_column_name] / merged_df['NATIONAL_POPULATION']) * 100
            prop_compared_prop_pop = f"{region}_PROPORTION_OF_NATIONAL_COMPARED_TO_PROP_POP"
            merged_df[prop_compared_prop_pop] = (merged_df[proportion_of_national_column] / merged_df[prop_national_pop_column]) * 100
        numeric_cols = [col for col in merged_df.columns if merged_df[col].dtype in ['int64', 'float64']]
        merged_df[numeric_cols] = merged_df[numeric_cols].round(1)
        SHS_with_population_calcs[df_name] = merged_df
        merged_df.to_csv(f'DATA/PROCESSED DATA/SHS/WithPopulation/{df_name}_WithPopulation.csv', index=False)
    return SHS_with_population_calcs

def long_formSHS(SHS_with_population_calcs, source_file):
    long_form_dfs = {}
    latest_dates = []
    for df_name, df in SHS_with_population_calcs.items():
        id_vars = ['DATE'] + [col for col in df.columns if df[col].dtype == 'object']
        value_vars = [col for col in df.columns if df[col].dtype in ['int64', 'float64']]
        long_form_dfs[df_name] = pd.melt(df, id_vars=id_vars, value_vars=value_vars, var_name='MEASURE', value_name='VALUE')
        long_form_dfs[df_name]['MEASURE'] = long_form_dfs[df_name]['MEASURE'].str.replace('_', ' ')
        long_form_dfs[df_name]['MEASURE'] = long_form_dfs[df_name]['MEASURE'].str.lower()
        long_form_dfs[df_name]['MEASURE'] = long_form_dfs[df_name]['MEASURE'].str.capitalize()
        #create column State, which is measure before first space
        long_form_dfs[df_name]['STATE'] = long_form_dfs[df_name]['MEASURE'].str.split(' ').str[0]
        #create column Measure, which is remaining measure after moving State to its own column
        long_form_dfs[df_name]['MEASURE'] = long_form_dfs[df_name]['MEASURE'].str.split(' ').str[1:].str.join(' ')
        long_form_dfs[df_name]['STATE'] = long_form_dfs[df_name]['STATE'].str.replace('Wa', 'WA').str.replace('Nsw', 'NSW').str.replace('Sa', 'SA').str.replace('Nt', 'NT').str.replace('Act', 'ACT')
        #move State column to second column
        cols = list(long_form_dfs[df_name].columns)
        cols.insert(1, cols.pop(cols.index('STATE')))
        long_form_dfs[df_name] = long_form_dfs[df_name][cols]
        long_form_dfs[df_name].to_csv(f'DATA/PROCESSED DATA/SHS/Long_Form/{df_name}_Long_Form.csv', index=False)
        latest_date = df['DATE'].max()
        latest_date = pd.to_datetime(latest_date)
        latest_dates.append(latest_date)


    latest_date = max(latest_dates)
    latest_date = pd.to_datetime(latest_date)
    update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
    update_log(latest_date, update_date, dataset= 'Monthly SHS data from AIHW')

    return 

def import_shs_data(source_file):
    path_to_dir = "DATA/PROCESSED DATA/SHS"
    prefix = 'SHS_'
    suffix = '.csv'
    Population_Sex_Age = pd.read_csv('DATA\PROCESSED DATA\Population\Population_State_Sex_Age_to_65+.csv')
    Population_Sex = pd.read_csv('DATA\PROCESSED DATA\Population\Population_State_Sex_Total.csv')
    Population_Total = pd.read_csv('DATA\PROCESSED DATA\Population\Population_State_Total_monthly.csv')
    get_SHS(source_file)
    processsed_dataframes = load_and_preprocess_data(path_to_dir, prefix, suffix)
    SHS_with_population_calcs = merge_and_calculate(processsed_dataframes, Population_Sex_Age, Population_Sex, Population_Total)
    long_formSHS(SHS_with_population_calcs, source_file)
    return

def get_airbnb(df, df_name):
    dfs = {}
    df_summaries = {}
    dfs[df_name] = df
    for df_name, df in dfs.items():
        df_summary_name = f"{df_name}_summary"
        df = df.groupby(['neighbourhood', 'room_type']).agg({'id': 'count', 'price': ['mean', 'median'], 'availability_365': ['mean', 'median']})
        df.columns = ['_'.join(col) for col in df.columns]
        df = df.reset_index()
        df = df.rename(columns={'id_count': 'count_listings'})
        df['date'] = df_name
        df_summaries[df_summary_name] = df

    df_summary = pd.concat(df_summaries.values())
    latest_date = df_summary['date'].max()
    latest_date = pd.to_datetime(latest_date).strftime('%d/%m/%Y')
    
    try:
        airbnb0 = pd.read_csv('DATA/PROCESSED DATA/Airbnb/airbnb_summary.csv')
        airbnb0 = pd.concat([airbnb0, df_summary])
        airbnb0 = airbnb0.drop_duplicates()
        airbnb0.to_csv('DATA/PROCESSED DATA/Airbnb/airbnb_summary.csv', index=False)
        update_log(latest_date, pd.to_datetime('today'), 'Airbnb')
    except:
        df_summary.to_csv('DATA/PROCESSED DATA/Airbnb/airbnb_summary.csv', index=False)
        update_log(latest_date, pd.to_datetime('today'), 'Airbnb')
    return

def state_total(df, df_name):
    filenames = os.listdir('DATA/SOURCE DATA/Airbnb')
    airbnbwa0 = pd.read_csv('DATA/PROCESSED DATA/Airbnb/Airbnb_WAtotals.csv')
    dfs = {}
    df_summaries = {}

    df['date'] = df_name
    dfs[df_name] = df
    all_details = pd.concat(dfs.values())
    for df_name, df in dfs.items():
        df_summary_name = f"{df_name}_summary"
        df = df.groupby(['room_type']).agg({'id': 'count', 'price': ['mean', 'median'], 'availability_365': ['mean', 'median']})
        df.columns = ['_'.join(col) for col in df.columns]
        df = df.reset_index()
        df = df.rename(columns={'id_count': 'count_listings'})
        df['date'] = df_name
        df_summaries[df_summary_name] = df
    df_summary_wa = pd.concat(df_summaries.values())
    
    try:
        df_summary_wa = pd.concat([airbnbwa0, df_summary_wa])
    except:
        pass

    df_summary_wa.to_csv('DATA/PROCESSED DATA/Airbnb_WAtotals.csv', index=False)
    all_details.to_csv('DATA/PROCESSED DATA/Airbnb_full.csv', index=False)
    return

def locs():
    df = pd.read_csv('DATA/PROCESSED DATA/Airbnb/airbnb_summary.csv')
    locs = pd.read_csv('DATA/Data descriptions/australian_postcodes.csv')
    #filter locs to WA
    locs = locs[locs['state'] == 'WA']
    #drop any sa4name = Northern Territory - Outback
    locs = locs[locs['sa4name'] != 'Northern Territory - Outback']
    #drop locs columns id, dc, type, state, status, sa3, sa4, region, SA1_MAINCODE_2011,	SA1_MAINCODE_2016,	SA2_MAINCODE_2016, SA3_CODE_2016, SA4_CODE_2016,	RA_2011	RA_2016	MMM_2015	MMM_2019	ced	altitude	chargezone	phn_code	phn_name
    locs = locs.drop(columns=['id', 'dc', 'type', 'state', 'status', 'sa3', 'sa4', 'sa3name', 'sa4name', 'region', 'SA1_MAINCODE_2011',	'SA1_MAINCODE_2016',	'SA2_MAINCODE_2016', 'SA3_CODE_2016', 'SA4_CODE_2016',	'RA_2011',	'RA_2016',	'MMM_2015',	'MMM_2019',	'altitude',	'chargezone',	'phn_code', 'long', 'lat', 'Lat_precise', 'Long_precise'])
    map = pd.read_csv('DATA/PROCESSED DATA/Airbnb/Airbnb_map.csv')
    map_old = map['old'].unique()


    df_full = pd.read_csv('DATA/PROCESSED DATA/Airbnb/Airbnb_summary.csv')
    if df_full['neighbourhood'].isin(map_old).any():
        df_full['neighbourhood'] = df_full['neighbourhood'].replace(map_old, map['new'])

    df_full = pd.merge(df_full, locs, left_on='neighbourhood', right_on='locality', how='left')

    if df['neighbourhood'].isin(map_old).any():
        df['neighbourhood'] = df['neighbourhood'].replace(map_old, map['new'])
    df = pd.merge(df, locs, left_on='neighbourhood', right_on='locality', how='left')

    df_full.to_csv('DATA/PROCESSED DATA/Airbnb/Airbnb_full.csv', index=False)
    df.to_csv('DATA/PROCESSED DATA/Airbnb/airbnb_summary.csv', index=False)
    return

def getPopulation():
    method = "get"
    url = "https://api.data.abs.gov.au/data/ABS,ERP_Q,1.0.0/1.2+1+3.A80+A75+A70+A65+A60+A55+A50+A45+A40+A35+A25+A30+A20+A15+A10+A59+A04+TOT..Q?startPeriod=2011-Q1"
    auth_string = f"{'x-api-key'}:{st.secrets['abskey']}"
    auth_string = auth_string.encode("ascii")
    auth_string = base64.b64encode(auth_string)
    headers = {
        'Accept': 'application/vnd.sdmx.data+csv;labels=both',
        'Authorization' : f"Basic {auth_string.decode('ascii')}"
    }
    response = requests.request(method, url, headers=headers, auth=None)

    if response.status_code == 200:
        content = response.content.decode('utf-8')
        csv_lines = content.splitlines()
        csv_reader = csv.reader(csv_lines)
        
        # Save the CSV content to a file
        with open(PopulationNewFile, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in csv_reader:
                csv_writer.writerow(row)
    return

def Waitlist_load_data(newWaitlistData):
    new_waitlist = pd.read_excel(newWaitlistData)
    current = pd.read_csv(SimpleWaitlistData)
    #
    df = pd.concat([current, new_waitlist])
    df = df.drop_duplicates()
    return df

def Waitlist_clear_new_data(newWaitlistData):
    wb = openpyxl.load_workbook(newWaitlistData)
    ws = wb.active
    ws.delete_rows(2, ws.max_row)
    wb.save(SimpleWaitlistData)
    return

def Waitlist_convert_to_long_form(df):
    df_long = df.melt(id_vars=["Date"], 
                      var_name="Category", 
                      value_name="Number")
    df_long['Date'] = pd.to_datetime(df_long['Date'], dayfirst=True)
    df_long = df_long.sort_values(['Category', 'Date'], ascending=[True, True])
    df_long = df_long.dropna(subset=['Number'])
    return df_long

def Waitlist_gap_filler(df_long):
    missing_dates = []
    Category_dfs = []
    for Category in df_long['Category'].unique():
        Category_df = df_long[df_long['Category'] == Category].copy()
        for i in range(len(Category_df)-1):
            if Category_df['Date'].iloc[i] + pd.DateOffset(days=1) + pd.offsets.MonthEnd(0) != Category_df['Date'].iloc[i+1]:    
                gap = round((Category_df['Date'].iloc[i+1] - Category_df['Date'].iloc[i]).days / 30) - 1
                diff = Category_df['Number'].iloc[i+1] - Category_df['Number'].iloc[i]
                for j in range(gap):
                    missing_date = Category_df['Date'].iloc[i] + pd.DateOffset(days=1) + pd.offsets.MonthEnd(0)
                    proxy_value = round(Category_df['Number'].iloc[i] + (diff / (gap+1)))
                    missing_dates.append(missing_date)
                    new_row = {'Date': missing_date, 'Category': Category, 'Number': proxy_value, 'Estimate flag - Number': 'Y'}
                    Category_df = pd.concat([Category_df, pd.DataFrame(new_row, index=[0])], ignore_index=True)
        Category_dfs.append(Category_df)
    df_long = pd.concat(Category_dfs)
    df_long = df_long.sort_values(by=['Date'])
    df_long = df_long.reset_index(drop=True)
    df_long['Estimate flag - Number'] = df_long['Estimate flag - Number'].fillna('')
    df_long['Estimate flag - Number'] = df_long['Estimate flag - Number'].astype(str)
    df_long['Date'] = pd.to_datetime(df_long['Date'])
    df_long['Category'] = df_long['Category'].astype(str)
    df_long['Number'] = df_long['Number'].astype(float)
    return df_long

def Waitlist_FYtdchange(df_long):
    for Category in df_long['Category'].unique():
        
        df_long_Category = df_long[df_long['Category'] == Category].copy()
        for i, row in df_long_Category.iterrows():
            if row['Date'].month > 6:
                eofy = str(row['Date'].year) + "-06-30"
            else:
                eofy = str((row['Date'].year)-1) + "-06-30"
            eofy = pd.to_datetime(eofy, format='%Y-%m-%d')
            if eofy in df_long_Category['Date'].values:
                df_long_Category.loc[i, 'Difference - financial year to date'] = row['Number'] - df_long_Category.loc[df_long_Category['Date']==eofy, 'Number'].values[0]
                df_long_Category.loc[i, 'Difference - financial year to date (per cent)'] = df_long_Category.loc[i, 'Difference - financial year to date'] / df_long_Category.loc[df_long_Category['Date']==eofy, 'Number'].values[0] * 100 
            else:
                df_long_Category.loc[i, 'Difference - financial year to date'] = float('nan')
                df_long_Category.loc[i, 'Difference - financial year to date (per cent)'] = float('nan')
        df_long = df_long.drop(df_long[df_long['Category'] == Category].index)
        df_long = pd.concat([df_long, df_long_Category])
    return df_long

def Waitlist_month_diff(df_long):
    df_long['Estimate flag - prior month'] = ''
    for Category in df_long['Category'].unique():
        df_long_Category = df_long[df_long['Category'] == Category].copy()
        df_long_Category['helper'] = df_long_Category['Date'] - pd.offsets.MonthEnd(1)
        df_long_Category['helper2'] = df_long_Category['helper'].apply(lambda x: df_long_Category.loc[df_long_Category['Date']==x, 'Number'].values[0] if x in df_long_Category['Date'].values else float('nan'))
        df_long_Category['Difference - prior month'] = df_long_Category['Number'] - df_long_Category['helper2']
        df_long_Category['Difference - prior month (per cent)'] = df_long_Category['Difference - prior month'] / df_long_Category['helper2'] * 100
        df_long_Category['Estimate flag - prior month'] = df_long_Category['helper2'].apply(lambda x: 'Y' if pd.isnull(x) else '')
        for i, row in df_long_Category.iterrows():
            if i == 0:
                continue
            if row['Estimate flag - prior month'] == 'Y':
                if i-1 in df_long_Category.index:
                    month_diff = row['Date'].month - df_long_Category.at[i-1, 'Date'].month
                    if month_diff != 0:
                        df_long_Category.loc[i, 'Difference -  month'] = (row['Number'] - df_long_Category.at[i-1, 'Number']) / month_diff
                        df_long_Category.loc[i, 'Difference - prior month (per cent)'] = df_long_Category.loc[i, 'Difference - prior month'] / df_long_Category.at[i-1, 'Number'] * 100
        df_long_Category = df_long_Category.drop(['helper', 'helper2'], axis=1)
        df_long = df_long.drop(df_long[df_long['Category'] == Category].index)
        df_long = pd.concat([df_long, df_long_Category])
    return df_long


def Waitlist_year_diff(df_long):
    df_long['Estimate flag - prior year'] = ''
    for Category in df_long['Category'].unique():
        df_long_Category = df_long[df_long['Category'] == Category].copy()
        df_long_Category['helper'] = df_long_Category['Date'] - pd.offsets.MonthEnd(12)
        df_long_Category['helper2'] = df_long_Category['helper'].apply(lambda x: df_long_Category.loc[df_long_Category['Date']==x, 'Number'].values[0] if x in df_long_Category['Date'].values else float('nan'))
        df_long_Category['Difference - prior year'] = df_long_Category['Number'] - df_long_Category['helper2']
        df_long_Category['Difference - prior year (per cent)'] = df_long_Category['Difference - prior year'] / df_long_Category['helper2'] * 100
        df_long_Category['Estimate flag - prior year'] = df_long_Category['helper2'].apply(lambda x: 'Y' if pd.isnull(x) else '')
        for i, row in df_long_Category.iterrows():
            if row['Date'] < df_long_Category['Date'].min() + pd.offsets.MonthEnd(12):
                continue
            if row['Estimate flag - prior year'] == 'Y':
                year_prior_date = row['Date'] - pd.offsets.MonthEnd(12)
                closest_date = min(df_long_Category['Date'], key=lambda x: abs(x - year_prior_date))
                df_long_Category.loc[i, 'Difference - prior year'] = (row['Number'] - df_long_Category.loc[df_long_Category['Date']==closest_date, 'Number'].values[0]/(row['Date'].month-closest_date.month/12))
                df_long_Category.loc[i, 'Difference - prior year (per cent)'] = df_long_Category.loc[i, 'Difference - prior year'] / df_long_Category.loc[df_long_Category['Date']==closest_date, 'Number'].values[0] * 100
        df_long_Category = df_long_Category.drop(['helper', 'helper2'], axis=1)
        df_long = df_long.drop(df_long[df_long['Category'] == Category].index)
        df_long = pd.concat([df_long, df_long_Category])
    return df_long

def Waitlist_calculate_12_month_average(df_long):
    def average_last_12_months(row):
        start_date = row['Date'] - pd.offsets.MonthEnd(11)
        end_date = row['Date']
        filtered_df = df_long[(df_long['Date'] >= start_date) & (df_long['Date'] <= end_date) & (df_long['Category'] == row['Category'])]
        return filtered_df['Number'].sum() / len(filtered_df)
    def average_prior_month_difference_last_12_months(row):
        start_date = row['Date'] - pd.offsets.MonthEnd(11)
        end_date = row['Date']
        filtered_df = df_long[(df_long['Date'] >= start_date) & (df_long['Date'] <= end_date) & (df_long['Category'] == row['Category'])]
        return filtered_df['Difference - prior month'].sum() / len(filtered_df)
    def average_prior_year_difference_last_12_months(row):
        start_date = row['Date'] - pd.offsets.MonthEnd(11)
        end_date = row['Date']
        filtered_df = df_long[(df_long['Date'] >= start_date) & (df_long['Date'] <= end_date) & (df_long['Category'] == row['Category'])]
        return filtered_df['Difference - prior year'].sum() / len(filtered_df)
    df_long['12 month rolling average'] = df_long.apply(average_last_12_months, axis=1)
    df_long['Difference - 12 month rolling average'] = df_long['Number'] - df_long['12 month rolling average']
    df_long['Difference - 12 month rolling average (per cent)'] = df_long['Difference - 12 month rolling average'] / df_long['12 month rolling average'] * 100
    df_long['12 month rolling average - prior month change'] = df_long.apply(average_prior_month_difference_last_12_months, axis=1)
    df_long['12 month rolling average - prior month change (per cent)'] = df_long['12 month rolling average - prior month change'] / df_long['12 month rolling average'] * 100
    df_long['12 month rolling average - prior year change'] = df_long.apply(average_prior_year_difference_last_12_months, axis=1)
    df_long['12 month rolling average - prior year change (per cent)'] = df_long['12 month rolling average - prior year change'] / df_long['12 month rolling average'] * 100
    return df_long

def Waitlist_calculate_cydiff(df_long):
    df_long['helper'] = df_long['Date'] - pd.offsets.YearEnd(1)
    df_long['helper2'] = df_long['helper'].apply(lambda x: df_long['Number'][df_long['Date']==x].values[0] if x in df_long['Date'].values else float('nan'))
    df_long['Difference - calendar year to date'] = df_long['Number'] - df_long['helper2']
    df_long['Difference - calendar year to date (per cent)'] = df_long['Difference - calendar year to date'] / df_long['helper2'] * 100
    df_long = df_long.drop(['helper', 'helper2'], axis=1)
    return df_long

def Waitlist_calculate_Priority_proportion(df_long):
    new_rows = pd.DataFrame()
    for date in df_long['Date'].unique():
        df_long_date = df_long[df_long['Date'] == date].copy()
        totalind = df_long_date[df_long_date['Category'] == 'Total individuals']['Number']
        priorityind = df_long_date[df_long_date['Category'] == 'Priority individuals']['Number']
        if totalind.empty or priorityind.empty:
            continue
        else:
            totalind = df_long_date[df_long_date['Category'] == 'Total individuals']['Number'].values[0]
            priorityind = df_long_date[df_long_date['Category'] == 'Priority individuals']['Number'].values[0]
            new_row_ind = pd.DataFrame([{'Date': date, 'Category': 'Proportion priority - individuals', 'Number': priorityind / totalind * 100}])
            new_rows = pd.concat([new_rows, new_row_ind])

            totalapp = df_long_date[df_long_date['Category'] == 'Total applications']['Number'].values[0]
            priorityapp = df_long_date[df_long_date['Category'] == 'Priority applications']['Number'].values[0]
            
            new_row_app = pd.DataFrame([{'Date': date, 'Category': 'Proportion priority - applications', 'Number': priorityapp / totalapp * 100}])
            new_rows = pd.concat([new_rows, new_row_app])

            new_row_avg_total = pd.DataFrame([{'Date': date, 'Category': 'Average number of individuals per application - total', 'Number': totalind / totalapp}])
            new_rows = pd.concat([new_rows, new_row_avg_total])

            new_row_avg_priority = pd.DataFrame([{'Date': date, 'Category': 'Average number of individuals per application - priority', 'Number': priorityind / priorityapp}])
            new_rows = pd.concat([new_rows, new_row_avg_priority])

            new_row_avg_nonpriority = pd.DataFrame([{'Date': date, 'Category': 'Average number of individuals per application - nonpriority', 'Number': (totalind - priorityind) / (totalapp - priorityapp)}])
            new_rows = pd.concat([new_rows, new_row_avg_nonpriority])

    df_long = pd.concat([df_long, new_rows])

    return df_long



def Waitlist_date_to_quarter_end(date):
    if date.month in [1, 2]:
        return pd.Timestamp(f'31-12-{(date.year)-1}')
    elif date.month in [4, 5]:
        return pd.Timestamp(f'31-03-{date.year}')
    elif date.month in [7, 8]:
        return pd.Timestamp(f'30-06-{date.year}')
    elif date.month in [10, 11]:
        return pd.Timestamp(f'30-09-{date.year}')
    else:
        return date

def Waitlist_add_quarter(df_long):
    df_long['Quarter'] = df_long['Date'].apply(Waitlist_date_to_quarter_end)
    return df_long



def Waitlist_nonpriority(df_long):
    new_rows = pd.DataFrame()
    for date in df_long['Date'].unique():
        df_long_date = df_long[df_long['Date'] == date].copy()
        totalind = df_long_date[df_long_date['Category'] == 'Total individuals']['Number']
        priorityind = df_long_date[df_long_date['Category'] == 'Priority individuals']['Number']
        if totalind.empty or priorityind.empty:
            continue
        else:
            totalind = df_long_date[df_long_date['Category'] == 'Total individuals']['Number'].values[0]
            priorityind = df_long_date[df_long_date['Category'] == 'Priority individuals']['Number'].values[0]
            new_row_ind = pd.DataFrame([{'Date': date, 'Category': 'nonPriority individuals', 'Number': totalind - priorityind}])
            new_rows = pd.concat([new_rows, new_row_ind])

        totalapp = df_long_date[df_long_date['Category'] == 'Total applications']['Number']
        priorityapp = df_long_date[df_long_date['Category'] == 'Priority applications']['Number']
        if totalapp.empty or priorityapp.empty:
            continue
        else:
            totalapp = df_long_date[df_long_date['Category'] == 'Total applications']['Number'].values[0]
            priorityapp = df_long_date[df_long_date['Category'] == 'Priority applications']['Number'].values[0]
            new_row_app = pd.DataFrame([{'Date': date, 'Category': 'nonPriority applications', 'Number': totalapp - priorityapp}])
            new_rows = pd.concat([new_rows, new_row_app])
    df_long = pd.concat([df_long, new_rows])
    return df_long

         
def Waitlist_add_population(df_long, population):
    population = population[['Date', 'WA_Population']]
    df_long = df_long.merge(population, how='left', left_on='Date', right_on='DATE')
    df_long['Percentage of population'] = df_long.apply(lambda row: row['Number'] / row['POPULATION'] * 100 if row['Category'] in ['Total individuals', 'Priority individuals', 'nonPriority individuals'] else float('nan'), axis=1)
    df_long['Value per 10 000'] = df_long.apply(lambda row: row['Number'] / row['POPULATION'] * 10000 if row['Category'] in ['Total individuals', 'Priority individuals', 'nonPriority individuals'] else float('nan'), axis=1)
    if 'Difference - prior month' in df_long.columns:
        df_long['Difference - prior month per 10 000'] = df_long.apply(lambda row: row['Difference - prior month'] / row['POPULATION'] * 10000 if row['Category'] in ['Total individuals', 'Priority individuals', 'nonPriority individuals'] else float('nan'), axis=1)
    if 'Difference - prior year' in df_long.columns:
        df_long['Difference - prior year per 10 000'] = df_long.apply(lambda row: row['Difference - prior year'] / row['POPULATION'] * 10000 if row['Category'] in ['Total individuals', 'Priority individuals', 'nonPriority individuals'] else float('nan'), axis=1)
    if '12 month rolling average' in df_long.columns:
        df_long['12 month rolling average per 10 000'] = df_long.apply(lambda row: row['12 month rolling average'] / row['POPULATION'] * 10000 if row['Category'] in ['Total individuals', 'Priority individuals', 'nonPriority individuals'] else float('nan'), axis=1)
    df_long = df_long.drop(['DATE'], axis=1)
    return df_long

def Waitlist_final_long(df_long, Waitlist_trend_longdf, Waitlist_latestdf):
    df_long['Category'] = df_long['Category'].str.replace('_', ' ')
    df_long['Category'] = df_long['Category'].str.title()
    df_long.loc[df_long['Category'].str.contains('Proportion'), 'Group'] = df_long.loc[df_long['Category'].str.contains('Proportion'), 'Category'].str.split(' - ').str[0]  
    df_long.loc[df_long['Category'].str.contains('Proportion'), 'Count'] = df_long.loc[df_long['Category'].str.contains('Proportion'), 'Category'].str.split(' - ').str[1]
    df_long.loc[df_long['Category'].str.contains('Average'), 'Group'] = df_long.loc[df_long['Category'].str.contains('Average'), 'Category'].str.split(' - ').str[0]
    df_long.loc[df_long['Category'].str.contains('Average'), 'Count'] = df_long.loc[df_long['Category'].str.contains('Average'), 'Category'].str.split(' - ').str[1]
    df_long.loc[df_long['Group'].isnull(), 'Group'] = df_long.loc[df_long['Group'].isnull(), 'Category'].str.split(' ').str[0]
    df_long.loc[df_long['Count'].isnull(), 'Count'] = df_long.loc[df_long['Count'].isnull(), 'Category'].str.split(' ').str[1]
    cols = df_long.columns.tolist()
    ids =["Date", "Category", "Group", "Count", "Estimate flag - Number", "Estimate flag - prior month", "Estimate flag - prior year"]
    values = [col for col in cols if col not in ids]
    df_long = df_long.melt(id_vars=ids, value_vars=values, var_name="Metric", value_name="Value")
    df_long = df_long[df_long['Metric'] != 'POPULATION']
    df_long['change from prior month'] = df_long.groupby(['Category', 'Metric'])['Value'].diff()
    df_long['change from prior month (per cent)'] = df_long['change from prior month'] / df_long['Value'].shift(1) * 100
    df_long['change from prior year'] = df_long.groupby(['Category', 'Metric'])['Value'].diff(12)
    df_long['change from prior year (per cent)'] = df_long['change from prior year'] / df_long['Value'].shift(12) * 100
    df_long.loc[(df_long['Metric'] == 'Difference - financial year to date') & (df_long['Date'].dt.month == 7), 'change from prior month'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - financial year to date (per cent)') & (df_long['Date'].dt.month == 7), 'change from prior month'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - calendar year to date') & (df_long['Date'].dt.month == 1), 'change from prior month'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - calendar year to date (per cent)') & (df_long['Date'].dt.month == 1), 'change from prior month'] = float('nan')
    df_long.loc[df_long['Metric'] == 'Number', 'change from prior month'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - financial year to date') & (df_long['Date'].dt.month == 7), 'change from prior year'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - financial year to date (per cent)') & (df_long['Date'].dt.month == 7), 'change from prior year'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - calendar year to date') & (df_long['Date'].dt.month == 1), 'change from prior year'] = float('nan')
    df_long.loc[(df_long['Metric'] == 'Difference - calendar year to date (per cent)') & (df_long['Date'].dt.month == 1), 'change from prior year'] = float('nan')
    df_long.loc[df_long['Metric'] == 'Number', 'change from prior year'] = float('nan')
    cols = df_long.columns.tolist()
    ids =["Date", "Category", "Group", "Count", "Estimate flag - Number", "Estimate flag - prior month", "Estimate flag - prior year", "Metric"]
    values = [col for col in cols if col not in ids]
    df_long = df_long.melt(id_vars=ids, value_vars=values, var_name="Value type", value_name="Number")
    df_long = df_long.dropna(subset=['Number'])
    df_long = df_long.rename(columns={'Number': 'Value'})


    df_long['Number type 1'] = df_long['Value']
    df_long['Number type 2'] = df_long['Value']
    df_long['Metric 1'] = df_long['Metric']
    df_long['Metric 2'] = df_long['Metric']

    df_long.loc[df_long['Metric'].str.contains('per cent'), 'Number type 1'] = 'Percentage'
    df_long.loc[~df_long['Metric'].str.contains('per cent'), 'Number type 1'] = 'Actual'
    df_long.loc[df_long['Metric'].str.contains('per cent'), 'Metric'] = df_long.loc[df_long['Metric'].str.contains('per cent'), 'Metric'].str.replace(' (per cent)', '')
    df_long.loc[df_long['Metric'].str.contains('Percentage'), 'Number type 1'] = 'Percentage'
    df_long.loc[df_long['Metric'].str.contains('per 10 000'), 'Number type 1'] = 'per 10 000'
    df_long.loc[df_long['Metric'].str.contains('per 10 000'), 'Metric'] = df_long.loc[df_long['Metric'].str.contains('per 10 000'), 'Metric'].str.replace(' per 10 000', '')
    df_long.loc[df_long['Value type'].str.contains('per cent'), 'Number type 2'] = 'Percentage'
    df_long.loc[~df_long['Value type'].str.contains('per cent'), 'Number type 2'] = 'Actual'
    df_long.loc[df_long['Value type'].str.contains('per cent'), 'Value type'] = df_long.loc[df_long['Value type'].str.contains('per cent'), 'Value type'].str.replace(' (per cent)', '')

    df_long.loc[df_long['Metric'].str.contains(' - '), 'Metric 1'] = df_long.loc[df_long['Metric'].str.contains(' - '), 'Metric'].str.split(' - ').str[0]
    df_long.loc[df_long['Metric'].str.contains(' - '), 'Metric 2'] = df_long.loc[df_long['Metric'].str.contains(' - '), 'Metric'].str.split(' - ').str[1]
    df_long.loc[~df_long['Metric'].str.contains(' - '), 'Metric 1'] = df_long.loc[~df_long['Metric'].str.contains(' - '), 'Metric']
    df_long.loc[~df_long['Metric'].str.contains(' - '), 'Metric 2'] = ''
    df_long = df_long.drop(['Metric'], axis=1)

    df_long['Estimate'] = 'N'
    
    df_long.loc[(df_long['Estimate flag - prior month'] == 'Y') & (df_long['Metric 1'].str.contains('prior month') | df_long['Metric 2'].str.contains('prior month') | df_long['Value type'].str.contains('prior month')), 'Estimate'] = 'Y'
    df_long.loc[(df_long['Estimate flag - prior year'] == 'Y') & (df_long['Metric 1'].str.contains('prior year') | df_long['Metric 2'].str.contains('prior year') | df_long['Value type'].str.contains('prior year')), 'Estimate'] = 'Y'
    df_long.loc[(df_long['Estimate flag - Number'] == 'Y') & (df_long['Metric 1'] == 'Number'), 'Estimate'] = 'Y'
    df_long = df_long.drop(['Estimate flag - prior month', 'Estimate flag - prior year', 'Estimate flag - Number'], axis=1)

    df_long = df_long.rename(columns={'Group': 'Description1', 'Count': 'Description2', 'Metric 1': 'Description3', 'Metric 2': 'Description4', 'Number type 1': 'Description5', 'Value type': 'Description6', 'Number type 2': 'Description7'})
    df_long = df_long[['Date', 'Category', 'Description1', 'Description2', 'Description3', 'Description4', 'Description5', 'Description6', 'Description7', 'Estimate', 'Value']]
    df_long.loc[df_long['Description3'] == 'Value', 'Description3'] = 'Number'
    df_long.loc[df_long['Description4'] == '', 'Description4'] = '-'
    df_long.loc[df_long['Description6'] == 'Value', 'Description7'] = '-'
    df_long.loc[df_long['Description6'] == 'Value', 'Description6'] = '-'
    df_long.to_csv(Waitlist_trend_longdf, index=False)
    #add column added_date, set to now datetime
    df_long['added_datetime'] = pd.to_datetime('now')



    df_latest = pd.DataFrame()
    for Category in df_long['Category'].unique():
        max_date = df_long[df_long['Category'] == Category]['Date'].max()
        df_cat_latest = df_long[(df_long['Category'] == Category) & (df_long['Date'] == max_date)]
        df_latest = pd.concat([df_latest, df_cat_latest])
    df_latest = df_latest.reset_index(drop=True)
    df_latest.to_csv(Waitlist_latestdf, index=False)
    return max_date


def import_waitlist_data():
    try:
        df = Waitlist_load_data(newWaitlistData)
        df_long = Waitlist_convert_to_long_form(df)
        df_long = Waitlist_gap_filler(df_long)
        df_long = Waitlist_nonpriority(df_long)
        df_long = Waitlist_calculate_Priority_proportion(df_long)
        df_long = Waitlist_add_population(df_long, pd.read_csv(PopulationStateMonthlydf))
        df_long = Waitlist_month_diff(df_long)
        df_long = Waitlist_year_diff(df_long)
        df_long = Waitlist_calculate_cydiff(df_long)
        df_long = Waitlist_calculate_12_month_average(df_long)
        df_long = Waitlist_FYtdchange(df_long)
        max_date = Waitlist_final_long(df_long, Waitlist_latestdf, Waitlist_trend_longdf)
        update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
        Waitlist_clear_new_data(newWaitlistData)
        update_log(max_date, update_date, 'Waitlist trend - statewide')
    except:
        pass
    return

def import_waitlist_manual(manualdata):
    df_long = Waitlist_convert_to_long_form(manualdata)
    df_long = Waitlist_gap_filler(df_long)
    df_long = Waitlist_nonpriority(df_long)
    df_long = Waitlist_calculate_Priority_proportion(df_long)
    df_long = Waitlist_add_population(df_long, pd.read_csv(PopulationStateMonthlydf))
    df_long = Waitlist_month_diff(df_long)
    df_long = Waitlist_year_diff(df_long)
    df_long = Waitlist_calculate_cydiff(df_long)
    df_long = Waitlist_calculate_12_month_average(df_long)
    df_long = Waitlist_FYtdchange(df_long)
    max_date = Waitlist_final_long(df_long, Waitlist_latestdf, Waitlist_trend_longdf)
    update_date = pd.to_datetime('today').strftime('%d/%m/%Y')
    Waitlist_clear_new_data(newWaitlistData)
    update_log(max_date, update_date, 'Waitlist trend - statewide')
    return

def census(census_data):
    if census_data == 'Total by state':
        df = pd.read_csv(ROGSHomelessnessdf, encoding='latin-1')
        df = df.sort_values(by='Year', ascending=True)
        df['Year'] = df['Year'].astype(str)
        df = df.rename(columns={'Aust': 'National'})
        regions = ['National', 'WA', 'Vic', 'Qld', 'SA', 'NSW', 'Tas', 'NT', 'ACT']

        df = df[df['Measure']=="Homelessness; by homelessness operational group"]

        df['Description2'] = df['Description2'].fillna('Persons')
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            Desc2 = st.selectbox('Select Description2 filter', df['Description2'].unique(), index=0)
            df = df[df['Description2'] == Desc2]
        with col2:
            if len(df['Description3'].unique()) > 1:
                Desc3 = st.selectbox('Select Description3 filter', df['Description3'].unique(), index=0)
                df = df[df['Description3'] == Desc3]
        with col3:
            if len(df['Description4'].unique()) > 1:
                Desc4 = st.selectbox('Select Description4 filter', df['Description4'].unique(), index=0)
                df = df[df['Description4'] == Desc4]
        with col4:
            datalabels = st.radio('Data labels', ['On', 'Off'], index=0, horizontal=True, key='censustotaldatalabels')

        df = df.sort_values(by=['Year'], ascending=True)
        #for region in regions, filter df for region, plotly bar, x=Year, y=Value, color=Region, group
        fig = go.Figure()
        yunits = df['Unit'].unique()[0]
        for region in regions:
            fig.add_trace(go.Bar(x=df['Year'], y=df[region], name=region))
        fig.update_layout(barmode='group', title='Homelessness; by homelessness operational group', xaxis_title="Year", yaxis_title=yunits)
        if datalabels == 'On':
            fig.update_traces(texttemplate='%{text:.2s}', textposition='inside', text=df[region])
        st.plotly_chart(fig, use_container_width=True)

    elif census_data == 'Geographic breakdown':
        st.markdown(f'Note - error in ABS table builder for LGAs, awaiting help on file and limited to SA4 for time being however quick swap in once received')
        df = pd.read_csv('DATA/PROCESSED DATA/Census/Multiyear/SA4Income_1621.csv')
        df = df.melt(id_vars=['OPGP HOMELESSNESS OPERATIONAL GROUPS', 'SA4', 'CENSUS_YEAR'], var_name='Income', value_name='Value')
        #replace Not Applicable with Not applicable
        df['OPGP HOMELESSNESS OPERATIONAL GROUPS'] = df['OPGP HOMELESSNESS OPERATIONAL GROUPS'].replace('Not Applicable', 'Not applicable')
        df = df[df['Income']=='TOTAL']
        df = df.drop(columns=['Income'])
        #in SA4, replace "Western Australia - " with ""
        df['SA4'] = df['SA4'].str.replace('Western Australia - ', '')
        #drop rows where Value = 0
        df = df[df['Value']!=0]
        col1, col2, COL3 = st.columns(3)
        with col1:
            geoselect = st.multiselect('Select area', df['SA4'].unique(), default = ['Perth - South East', 'Perth - North West', 'Perth - North East', 'Perth - South West', 'Perth - Inner'])
        with col2:
            datalabels = st.radio('Data labels', ['On', 'Off'], index=0, horizontal=True, key='censusSA4datalabels')
        dftotal = df[df['OPGP HOMELESSNESS OPERATIONAL GROUPS']=='Total']
        dfna = df[df['OPGP HOMELESSNESS OPERATIONAL GROUPS']=='Not applicable']
        dftotalna = pd.concat([dftotal, dfna])
        #pivot so OPGP HOMELESSNESS OPERATIONAL GROUPS is columns
        dftotalna = dftotalna.pivot(index=['SA4', 'CENSUS_YEAR'], columns='OPGP HOMELESSNESS OPERATIONAL GROUPS', values='Value').reset_index()
        dftotalna['Total homelessness'] = dftotalna['Total'] - dftotalna['Not applicable']
        dftotalna['per 10k'] = dftotalna['Total homelessness'] / dftotalna['Total'] * 10000
        #drop Not applicable
        dftotalna = dftotalna.drop(columns=['Not applicable'])
        #melt back to long format
        dftotalna = dftotalna.melt(id_vars=['SA4', 'CENSUS_YEAR'], var_name='Measure', value_name='Value')
        df = df[df['OPGP HOMELESSNESS OPERATIONAL GROUPS']!='Total']
        df = df[df['OPGP HOMELESSNESS OPERATIONAL GROUPS']!='Not applicable']
        df16 = df[df['CENSUS_YEAR']==2016]
        df21 = df[df['CENSUS_YEAR']==2021]

        #filter dftotal21 for geoselect
        df10k = dftotalna[dftotalna['SA4'].isin(geoselect)]
        df10k = df10k[df10k['Measure']=='per 10k']
        df10k = df10k.drop(columns=['Measure'])

        dftotal = dftotalna[dftotalna['SA4'].isin(geoselect)]
        dftotal = dftotal[dftotal['Measure']=='Total homelessness']
        dftotal = dftotal.drop(columns=['Measure'])

        dftotal16 = dftotal[dftotal['CENSUS_YEAR']==2016]
        dftotal21 = dftotal[dftotal['CENSUS_YEAR']==2021]

        fig = go.Figure()
        df10k16 = df10k[df10k['CENSUS_YEAR']==2016]
        df10k21 = df10k[df10k['CENSUS_YEAR']==2021]
        
        
        
        if datalabels == 'On':
            fig.add_trace(go.Bar(x=df10k21['SA4'], y=df10k21['Value'], name='2021', text=df10k21['Value'], textposition='inside', texttemplate='%{text:.2s}'))
            fig.add_trace(go.Bar(x=df10k16['SA4'], y=df10k16['Value'], name='2016', text=df10k16['Value'], textposition='inside', texttemplate='%{text:.2s}'))
        else:
            fig.add_trace(go.Bar(x=df10k21['SA4'], y=df10k21['Value'], name='2021'))
            fig.add_trace(go.Bar(x=df10k16['SA4'], y=df10k16['Value'], name='2016'))
        
        fig.update_layout(barmode='group', title='Homelessness; per 10k population', yaxis_title='Value', xaxis_tickangle=-45, height=800, width=3200/len(geoselect))
        st.plotly_chart(fig)

        fig = go.Figure()
        if datalabels == 'On':
            fig.add_trace(go.Bar(x=dftotal21['SA4'], y=dftotal21['Value'], name='2021', text=dftotal21['Value'], textposition='inside', texttemplate='%{text:.2s}'))
            fig.add_trace(go.Bar(x=dftotal16['SA4'], y=dftotal16['Value'], name='2016', text=dftotal16['Value'], textposition='inside', texttemplate='%{text:.2s}'))
        else:
            fig.add_trace(go.Bar(x=dftotal16['SA4'], y=dftotal16['Value'], name='2016'))
            fig.add_trace(go.Bar(x=dftotal21['SA4'], y=dftotal21['Value'], name='2021'))

        fig.update_layout(barmode='group', title='Homelessness; total', yaxis_title='Value', xaxis_tickangle=-45, height=800, width=3200/len(geoselect))
        st.plotly_chart(fig)

        for sa4 in geoselect:
            filtereddf16 = df16[df16['SA4']==sa4]
            filtereddf21 = df21[df21['SA4']==sa4]
        
            #plotly bar, x=OPGP HOMELESSNESS OPERATIONAL GROUPS, y=Value, color=CENSUS_YEAR
            fig = go.Figure()
            fig.add_trace(go.Bar(x=filtereddf16['OPGP HOMELESSNESS OPERATIONAL GROUPS'], y=filtereddf16['Value'], name='2016'))
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{text:.2s}', textposition='inside', text=filtereddf16['Value'])
            fig.add_trace(go.Bar(x=filtereddf21['OPGP HOMELESSNESS OPERATIONAL GROUPS'], y=filtereddf21['Value'], name='2021'))
            if datalabels == 'On':
                fig.update_traces(texttemplate='%{text:.2s}', textposition='inside', text=filtereddf21['Value'])
            fig.update_layout(barmode='group', title=f'Homelessness groups - {sa4}', yaxis_title='persons', height=800)
            st.plotly_chart(fig, use_container_width=True)

            #pie chart for each year
            col1, col2 = st.columns(2)
            
            fig = go.Figure()
            fig.add_trace(go.Pie(labels=filtereddf16['OPGP HOMELESSNESS OPERATIONAL GROUPS'], values=filtereddf16['Value'], name='2016'))
            fig.update_layout(title=f'Homelessness groups - 2016 - {sa4}', yaxis_title='persons')
            if datalabels == 'On':
                fig.update_traces(textposition='inside', textinfo='percent+value')
            with col1:
                st.plotly_chart(fig)
            fig = go.Figure()
            fig.add_trace(go.Pie(labels=filtereddf21['OPGP HOMELESSNESS OPERATIONAL GROUPS'], values=filtereddf21['Value'], name='2021'))
            fig.update_layout(title=f'Homelessness groups - 2021 - {sa4}', xaxis_title="Homelessness operational group", yaxis_title='Value')
            if datalabels == 'On':
                fig.update_traces(textposition='inside', textinfo='percent+value')
            with col2:    
                st.plotly_chart(fig)
    elif census_data == 'Aboriginal and Torres Strait Islander status':
        df = pd.read_csv('DATA/PROCESSED DATA/Census/Multiyear/ATSIState_1621.csv')
        df = df.melt(id_vars=['INGP INDIGENOUS STATUS', 'STATE', 'CENSUS_YEAR'], var_name='Group', value_name='Value')
        
        df['Group'] = df['Group'].str.replace('PERSONS IN', '')
        #replace PERSONS WHO ARE
        df['Group'] = df['Group'].str.replace('PERSONS WHO ARE', '')
        #replace PERSONS LIVING IN 
        df['Group'] = df['Group'].str.replace('PERSONS LIVING IN', '')
        #replace PERSONS 
        df['Group'] = df['Group'].str.replace('PERSONS', '')
        #lower case
        df['Group'] = df['Group'].str.lower()
        
        #drop INGP INDIGENOUS STATUS = Total
        df = df[df['INGP INDIGENOUS STATUS']!='Total']
        #IN state, replace Western Australia - with 'WA', South Australia - with 'SA', New South Wales - with 'NSW', Victoria - with 'Vic', Queensland - with 'Qld', Tasmania - with 'Tas', Northern Territory - with 'NT', Australian Capital Territory - with 'ACT'
        df['STATE'] = df['STATE'].str.replace('Western Australia', 'WA')
        df['STATE'] = df['STATE'].str.replace('South Australia', 'SA')
        df['STATE'] = df['STATE'].str.replace('New South Wales', 'NSW')
        df['STATE'] = df['STATE'].str.replace('Victoria', 'Vic')
        df['STATE'] = df['STATE'].str.replace('Queensland', 'Qld')
        df['STATE'] = df['STATE'].str.replace('Tasmania', 'Tas')
        df['STATE'] = df['STATE'].str.replace('Northern Territory', 'NT')
        df['STATE'] = df['STATE'].str.replace('Australian Capital Territory', 'ACT')

        #YEAR ASCENDING ORDER
        df = df.sort_values(by=['CENSUS_YEAR'], ascending=True)
        groupslist = df['Group'].unique()
        #remove TOTAL
        groupslist = groupslist[groupslist!='total']
        #remove Not applicable
        groupslist = groupslist[groupslist!='not applicable']
        group_select = st.multiselect('Select group', groupslist, default = groupslist)
        
        df2 = df[df['Group'].isin(group_select)]
        #select STATE
        col1, col2 = st.columns(2)
        with col1:
            state = st.selectbox('Select state', df2['STATE'].unique(), index=6)
        with col2:
            year = st.selectbox('Select year', df2['CENSUS_YEAR'].unique(), index=1)
        df2 = df2[df2['STATE']==state]
        df2 = df2[df2['CENSUS_YEAR']==year]
        fig = go.Figure()
        for group in df['Group'].unique():
            filtereddf = df2[df2['Group']==group]
            fig.add_trace(go.Bar(x=filtereddf['INGP INDIGENOUS STATUS'], y=filtereddf['Value'], name=group, text=filtereddf['Value'], textposition='inside', texttemplate='%{text:.2s}'))
        fig.update_layout(barmode='stack', title=f'Aboriginal and Torres Strait Islander status - {state}', yaxis_title='persons', height=800, xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)
        dftotal = df[df['Group']=='total']
        #filter year ascending
        dftotal = dftotal.sort_values(by=['CENSUS_YEAR'], ascending=True)
        #year as string
        dftotal['CENSUS_YEAR'] = dftotal['CENSUS_YEAR'].astype(str)
        selectyear = st.selectbox('Select year', dftotal['CENSUS_YEAR'].unique(), index=1, key='censusatsiselectyear2')
        dftotal = dftotal[dftotal['CENSUS_YEAR']==selectyear]
        fig = go.Figure()
        for atsistatus in dftotal['INGP INDIGENOUS STATUS'].unique():
            filtereddf = dftotal[dftotal['INGP INDIGENOUS STATUS']==atsistatus]
            fig.add_trace(go.Bar(x=filtereddf['STATE'], y=filtereddf['Value'], name=atsistatus, text=filtereddf['Value'], textposition='inside', texttemplate='%{text:.2s}'))
            fig.update_layout(barmode='stack', title=f'Aboriginal and Torres Strait Islander status - all homelessness groups', yaxis_title='persons', height=600, xaxis_tickangle=-45)
            
        #move xaxis label up
        fig.update_xaxes(title_standoff=0)
        st.plotly_chart(fig, use_container_width=True)
    return





if __name__ == "__main__":
    data_updates()
    home()

